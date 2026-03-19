"""
_update_cert_analysis.py  (v2)
==============================
Aggiorna CERT_COMPLETO applicando:
  1. Correzioni nome/cognome/email da PERSONE_SENZA_EMAIL (incluse inversioni)
  2. Colori:
       VERDE   = AI conferma persona + cert (AI_PERSONA=SI, AI_CERT=SI)
       ROSSO   = qualsiasi mismatch AI (persona=NO OR cert=NO) → riga da eliminare
       GRIGIO  = nessun PDF / non verificato
  3. Per ogni riga ROSSA: aggiunge una riga AZZURRA con i dati corretti estratti dall'AI
     (persona_trovata + cert_trovata + email risolta)

NON riesegue l'AI. Usa solo CERT_COMPLETO e AI_VERIFICA già presenti nel file.
"""

import io, os, re, sys, unicodedata
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ─── Percorsi ────────────────────────────────────────────────────────────────
BASE     = r"C:\Users\GiuseppeComparetti\MASHFROG GROUP S.R.L\ENT_SOLUTION_M4P_STAFF - Documenti"
XLS_FILE = os.path.join(BASE, "CERT_ANALYSIS_20260319.xlsx")
XLS_SRC  = os.path.join(BASE, "SAP_CERTIFICAZIONI_2026.xlsx")   # per email map Sheet2

# ─── Colori ──────────────────────────────────────────────────────────────────
FILL = {
    "VERDE":   PatternFill("solid", fgColor="C8E6C9"),
    "ROSSO":   PatternFill("solid", fgColor="FFCDD2"),
    "AZZURRO": PatternFill("solid", fgColor="BBDEFB"),
    "GRIGIO":  PatternFill("solid", fgColor="F5F5F5"),
    "SEPARAT": PatternFill("solid", fgColor="E3F2FD"),
    "HEADER":  PatternFill("solid", fgColor="1F4E79"),
}
FONT_HDR  = Font(bold=True, color="FFFFFF", size=10)
FONT_NEW  = Font(italic=True, color="1565C0")
FONT_SEP  = Font(bold=True, color="1565C0", size=11)
WRAP      = Alignment(wrap_text=True, vertical="top")

LEGENDA = [
    ("LEGENDA COLORI",                                    "1F4E79","FFFFFF", True),
    ("VERDE   — AI conferma persona + cert corretti",     "C8E6C9","1B5E20", False),
    ("ROSSO   — Mismatch AI: riga da eliminare",          "FFCDD2","B71C1C", False),
    ("AZZURRO — Riga corretta suggerita dall'AI",         "BBDEFB","1565C0", False),
    ("GRIGIO  — Nessun PDF / non verificato dall'AI",     "F5F5F5","616161", False),
]

# ─── Helper ───────────────────────────────────────────────────────────────────
def _norm(s):
    if not s: return ""
    s = unicodedata.normalize("NFD", str(s).strip())
    return "".join(c for c in s if unicodedata.category(c) != "Mn").lower()

def _clean(v):
    if v is None: return ""
    return str(v).strip().replace("\xa0","").strip()

def _sim(a, b):
    return SequenceMatcher(None, _norm(a), _norm(b)).ratio()

def _fill_row(ws, row_idx, fill, font=None):
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.alignment = WRAP
        if font: cell.font = font

# ═════════════════════════════════════════════════════════════════════════════
# 1. EMAIL MAP DA Sheet2 del SAP EXCEL ORIGINALE
# ═════════════════════════════════════════════════════════════════════════════
print("[1/5] Carico email map da SAP_CERTIFICAZIONI_2026.xlsx Sheet2 ...")
wb_src  = openpyxl.load_workbook(XLS_SRC, data_only=True)
ws_src2 = wb_src[wb_src.sheetnames[1]]
rows_s2 = list(ws_src2.iter_rows(values_only=True))

# {norm(cognome+nome): email}  e  {norm(cognome): [(cog,nom,email)]}
email_map   : dict[str, str]  = {}
surname_idx : dict[str, list] = defaultdict(list)

for r in rows_s2[1:]:
    emp   = _clean(r[4])   # Employee
    email = _clean(r[6]).lower()
    if not emp or "@" not in email: continue
    parts = emp.split()
    cog, nom = parts[0], " ".join(parts[1:])
    email_map[_norm(f"{cog} {nom}")] = email
    surname_idx[_norm(cog)].append((cog, nom, email))

print(f"   {len(email_map)} email caricate")

def _lookup_email(name_raw: str) -> tuple[str,str,str,float]:
    """Cerca email da nome libero. Ritorna (cog, nom, email, conf)."""
    name_raw = name_raw.strip()
    parts    = name_raw.split()
    if not parts: return "","","",0.0
    # prova entrambi gli ordini
    for order in [name_raw, " ".join(reversed(parts))]:
        k = _norm(order)
        if k in email_map:
            p = order.split()
            return p[0], " ".join(p[1:]), email_map[k], 1.0
    # fuzzy
    best_s, best_k, best_e = 0.0,"",""
    for k, em in email_map.items():
        s = max(_sim(name_raw, k), _sim(" ".join(reversed(parts)), k))
        if s > best_s: best_s,best_k,best_e = s,k,em
    if best_s >= 0.80:
        p = best_k.split()
        return p[0].capitalize()," ".join(p[1:]).capitalize(),best_e,round(best_s,3)
    # solo cognome
    for st in [parts[0], parts[-1]]:
        cands = surname_idx.get(_norm(st),[])
        if len(cands)==1:
            return cands[0][0],cands[0][1],cands[0][2],0.75
    return "","","",0.0

# ═════════════════════════════════════════════════════════════════════════════
# 2. MAPPA CORREZIONI DA PERSONE_SENZA_EMAIL
# ═════════════════════════════════════════════════════════════════════════════
print("[2/5] Costruisco mappa correzioni da PERSONE_SENZA_EMAIL ...")

wb_main = openpyxl.load_workbook(XLS_FILE)
ws_pse  = wb_main["PERSONE_SENZA_EMAIL"]
rows_pse = list(ws_pse.iter_rows(values_only=True))
# Headers: COGNOME, NOME, N_CERT, POSSIBILI_EMAIL, NOTE_REVISIONE, Nuova mail

# {norm(cog+" "+nom): {email, new_cog, new_nom}}
corr_map: dict[str, dict] = {}

for r in rows_pse[1:]:
    cog       = _clean(r[0])
    nom       = _clean(r[1])
    email_sug = _clean(r[3])
    note      = _clean(r[4]).lower()
    nuova     = _clean(r[5]) if len(r) > 5 else ""

    if not cog: continue

    email_final = nuova if nuova else email_sug

    # Determina nome/cognome corretti
    if "inverti" in note:
        # Il cog e nom sono scambiati → correggo
        new_cog, new_nom = nom, cog
        # Caso speciale: "Elisa De" / "Mattia" → correct: De Mattia / Elisa
        if _norm(nom).startswith("elisa de") or _norm(cog) == "mattia":
            new_cog, new_nom = "De Mattia", "Elisa"
    elif "nome:" in note and "cognome:" in note:
        # Es. "Nome: Annamaria Cognome: Suglia"
        m_n = re.search(r"nome:\s*([^\s,]+(?:\s+[^\s,]+)*?)(?:\s+cognome:|\s*$)", note, re.I)
        m_c = re.search(r"cognome:\s*([^\s,]+(?:\s+[^\s,]+)*?)(?:\s+nome:|\s*$)", note, re.I)
        new_nom = m_n.group(1).strip().capitalize() if m_n else nom
        new_cog = m_c.group(1).strip().capitalize() if m_c else cog
    else:
        new_cog, new_nom = cog, nom

    key = _norm(f"{cog} {nom}")
    corr_map[key] = {"email": email_final, "cog": new_cog, "nom": new_nom}
    # Aggiungi anche la chiave invertita per robustezza
    key2 = _norm(f"{nom} {cog}")
    if key2 != key:
        corr_map[key2] = {"email": email_final, "cog": new_cog, "nom": new_nom}

print(f"   {len(corr_map)} correzioni persona mappate")

def _apply_person_corr(cog: str, nom: str, email: str) -> tuple[str,str,str]:
    """Ritorna (cog_corretto, nom_corretto, email_corretta)."""
    k = _norm(f"{cog} {nom}")
    if k in corr_map:
        c = corr_map[k]
        return c["cog"], c["nom"], c["email"]
    # Prova solo cognome (per persone già con email ma con nome sbagliato)
    k2 = _norm(f"{nom} {cog}")
    if k2 in corr_map:
        c = corr_map[k2]
        return c["cog"], c["nom"], c["email"]
    return cog, nom, email

# ═════════════════════════════════════════════════════════════════════════════
# 3. LEGGI AI_VERIFICA → {stem_file: ai_result}
# ═════════════════════════════════════════════════════════════════════════════
print("[3/5] Carico AI_VERIFICA ...")
ws_ai     = wb_main["AI_VERIFICA"]
rows_ai   = list(ws_ai.iter_rows(values_only=True))
ai_hdrs   = [_clean(h) for h in rows_ai[0]]

def _aicol(n): return ai_hdrs.index(n)
AI_FILE=_aicol("FILE"); AI_MP=_aicol("MATCH_PERSONA"); AI_MC=_aicol("MATCH_CERT")
AI_PT=_aicol("PERSONA_TROVATA"); AI_CT=_aicol("CERT_TROVATA")
AI_AT=_aicol("ANNO_TROVATO"); AI_NOTE=_aicol("NOTE_AI"); AI_ERR=_aicol("ERRORE")
AI_CONF=_aicol("CONFIDENZA")

ai_by_stem: dict[str, dict] = {}
for r in rows_ai[1:]:
    fname = _clean(r[AI_FILE])
    if not fname: continue
    ai_by_stem[Path(fname).stem] = {
        "match_persona": _clean(r[AI_MP]),
        "match_cert":    _clean(r[AI_MC]),
        "persona_trovata": _clean(r[AI_PT]),
        "cert_trovata":    _clean(r[AI_CT]),
        "anno_trovato":    _clean(r[AI_AT]),
        "note_ai":         _clean(r[AI_NOTE]),
        "errore":          _clean(r[AI_ERR]),
        "confidenza":      _clean(r[AI_CONF]),
    }
print(f"   {len(ai_by_stem)} risultati AI")

# ═════════════════════════════════════════════════════════════════════════════
# 4. PROCESSA CERT_COMPLETO
# ═════════════════════════════════════════════════════════════════════════════
print("[4/5] Aggiorno CERT_COMPLETO ...")

ws_cert   = wb_main["CERT_COMPLETO"]
rows_cert = list(ws_cert.iter_rows(values_only=True))
hdrs_cert = [_clean(h) for h in rows_cert[0]]

def _cc(n):
    try: return hdrs_cert.index(n)
    except ValueError: return -1

CC = {n: _cc(n) for n in [
    "FONTE","COGNOME","NOME","EMAIL","EMAIL_CONF","CERT_CODE","CERT_NAME",
    "AREA","CLUSTER","ANNO","STATUS","PDF_FILE","PDF_CATEGORIA","PDF_CONF",
    "AI_PERSONA","AI_CERT","AI_CONFIDENZA","NOTE_ORIG","FLAG","NOTE_REVISIONE"
]}

n_cols = len(hdrs_cert)
stat   = {"verde":0,"rosso":0,"grigio":0,"corr_person":0}
new_rows: list[list] = []   # righe azzurre da aggiungere in fondo
processed_rows = set()      # righe già processate (evita doppi)

for row_i, row_data in enumerate(rows_cert[1:], start=2):

    # ── 4a. Applica correzioni nome/cognome/email ──────────────────────────
    cog   = _clean(row_data[CC["COGNOME"]])
    nom   = _clean(row_data[CC["NOME"]])
    email = _clean(row_data[CC["EMAIL"]])

    new_cog, new_nom, new_email = _apply_person_corr(cog, nom, email)

    if new_cog != cog or new_nom != nom or new_email != email:
        ws_cert.cell(row=row_i, column=CC["COGNOME"]+1).value = new_cog
        ws_cert.cell(row=row_i, column=CC["NOME"]+1).value    = new_nom
        ws_cert.cell(row=row_i, column=CC["EMAIL"]+1).value   = new_email
        if CC["EMAIL_CONF"] >= 0:
            ws_cert.cell(row=row_i, column=CC["EMAIL_CONF"]+1).value = "100%"
        # Rimuovi NO_EMAIL dal flag se email ora presente
        if new_email and CC["FLAG"] >= 0:
            old_flag = _clean(ws_cert.cell(row=row_i, column=CC["FLAG"]+1).value)
            new_flag = " | ".join(f.strip() for f in old_flag.split("|")
                                   if f.strip() and f.strip() != "NO_EMAIL")
            ws_cert.cell(row=row_i, column=CC["FLAG"]+1).value = new_flag
        stat["corr_person"] += 1
        cog, nom, email = new_cog, new_nom, new_email

    # ── 4b. Determina colore da AI_VERIFICA ───────────────────────────────
    pdf_file = _clean(row_data[CC["PDF_FILE"]]).replace(" (?)","").strip()

    if not pdf_file:
        _fill_row(ws_cert, row_i, FILL["GRIGIO"])
        stat["grigio"] += 1
        continue

    ai = ai_by_stem.get(Path(pdf_file).stem)
    if not ai:
        _fill_row(ws_cert, row_i, FILL["GRIGIO"])
        stat["grigio"] += 1
        continue

    mp = ai["match_persona"]
    mc = ai["match_cert"]

    is_mismatch = (mp == "NO" or mc == "NO")

    if not is_mismatch:
        # VERDE
        _fill_row(ws_cert, row_i, FILL["VERDE"])
        stat["verde"] += 1
        if CC["NOTE_REVISIONE"] >= 0:
            c = ws_cert.cell(row=row_i, column=CC["NOTE_REVISIONE"]+1)
            if not _clean(c.value):
                c.value = "✅ AI conferma persona e certificazione corrette"
        continue

    # ROSSO — mismatch
    _fill_row(ws_cert, row_i, FILL["ROSSO"])
    stat["rosso"] += 1

    # Nota nel campo revisione
    parts_note = []
    if mp == "NO":
        parts_note.append(f"Persona nel PDF: '{ai['persona_trovata']}'")
    if mc == "NO":
        parts_note.append(f"Cert nel PDF: '{ai['cert_trovata']}'")
    parts_note.append(ai["note_ai"])
    if CC["NOTE_REVISIONE"] >= 0:
        ws_cert.cell(row=row_i, column=CC["NOTE_REVISIONE"]+1).value = (
            "❌ MISMATCH — " + " | ".join(p for p in parts_note if p)
        )

    # ── 4c. Costruisci riga azzurra corretta ──────────────────────────────
    persona_raw  = ai["persona_trovata"]
    cert_trovata = ai["cert_trovata"]
    anno_trovato = ai["anno_trovato"]

    # Persona corretta
    if mp == "SI":
        # Persona ok → usa dati già corretti di questa riga
        az_cog, az_nom, az_email, az_econf = cog, nom, email, "100%"
    else:
        # Cerca persona trovata dall'AI
        az_cog, az_nom, az_email, az_econf_f = _lookup_email(persona_raw)
        az_econf = f"{az_econf_f:.0%}" if az_econf_f else "?"
        if not az_cog and persona_raw:
            # Fallback: usa il nome raw splittato
            p = persona_raw.split()
            az_cog = p[-1] if p else persona_raw
            az_nom = " ".join(p[:-1]) if len(p) > 1 else ""
        # Applica correzioni anche al nome trovato dall'AI
        az_cog, az_nom, az_email2 = _apply_person_corr(az_cog, az_nom, az_email)
        if az_email2: az_email = az_email2

    # Cert corretta
    if mc == "SI":
        az_code  = _clean(row_data[CC["CERT_CODE"]])
        az_cname = _clean(row_data[CC["CERT_NAME"]])
    else:
        az_code  = ""   # non conosciamo il codice della cert trovata nel PDF
        az_cname = cert_trovata

    az_anno = anno_trovato or _clean(row_data[CC["ANNO"]])

    note_az = (f"Riga suggerita dall'AI. "
               f"Persona nel PDF: '{persona_raw}'. "
               f"Cert nel PDF: '{cert_trovata}'. "
               f"Anno trovato: '{anno_trovato}'. "
               f"Note AI: {ai['note_ai']}")

    az_row = [""] * n_cols
    az_row[CC["FONTE"]]     = "AI_SUGGERITO"
    az_row[CC["COGNOME"]]   = az_cog
    az_row[CC["NOME"]]      = az_nom
    az_row[CC["EMAIL"]]     = az_email
    az_row[CC["EMAIL_CONF"]]= az_econf
    az_row[CC["CERT_CODE"]] = az_code
    az_row[CC["CERT_NAME"]] = az_cname
    az_row[CC["AREA"]]      = _clean(row_data[CC["AREA"]])
    az_row[CC["CLUSTER"]]   = _clean(row_data[CC["CLUSTER"]])
    az_row[CC["ANNO"]]      = az_anno
    az_row[CC["PDF_FILE"]]  = pdf_file
    az_row[CC["PDF_CATEGORIA"]] = _clean(row_data[CC["PDF_CATEGORIA"]])
    az_row[CC["AI_PERSONA"]]= "SI"
    az_row[CC["AI_CERT"]]   = "SI"
    az_row[CC["AI_CONFIDENZA"]] = ai["confidenza"]
    az_row[CC["FLAG"]]      = "AI_SUGGERITO"
    az_row[CC["NOTE_REVISIONE"]] = note_az

    new_rows.append(az_row)

# ═════════════════════════════════════════════════════════════════════════════
# 5. AGGIUNGI RIGHE AZZURRE IN FONDO
# ═════════════════════════════════════════════════════════════════════════════

if new_rows:
    # Separatore
    sep_i = ws_cert.max_row + 2
    sep_c = ws_cert.cell(row=sep_i, column=1,
        value="▼  RIGHE CORRETTE DALL'AI  — verificare email e cert_code, poi spostare nella posizione corretta ed eliminare le righe rosse")
    sep_c.fill = FILL["SEPARAT"]
    sep_c.font = FONT_SEP
    ws_cert.merge_cells(start_row=sep_i, start_column=1,
                        end_row=sep_i, end_column=min(n_cols, 20))

    for az_row in new_rows:
        r_i = ws_cert.max_row + 1
        for col_i, val in enumerate(az_row, start=1):
            c = ws_cert.cell(row=r_i, column=col_i, value=val)
            c.fill      = FILL["AZZURRO"]
            c.font      = FONT_NEW
            c.alignment = WRAP

# ── Legenda (colonne a destra) ─────────────────────────────────────────────
leg_col = n_cols + 2
for i, (text, bg, fg, bold) in enumerate(LEGENDA, start=1):
    c = ws_cert.cell(row=i, column=leg_col, value=text)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
ws_cert.column_dimensions[
    ws_cert.cell(1, leg_col).column_letter].width = 55

# ═════════════════════════════════════════════════════════════════════════════
# SALVA
# ═════════════════════════════════════════════════════════════════════════════
print("[5/5] Salvo ...")
wb_main.save(XLS_FILE)

print()
print("=" * 62)
print(f"  FILE AGGIORNATO: {Path(XLS_FILE).name}")
print("=" * 62)
print(f"  Correzioni persona applicate: {stat['corr_person']}")
print(f"  Righe VERDE  (AI OK):         {stat['verde']}")
print(f"  Righe ROSSO  (mismatch):      {stat['rosso']}")
print(f"  Righe GRIGIO (no PDF/no AI):  {stat['grigio']}")
print(f"  Righe AZZURRO generate:       {len(new_rows)}")
