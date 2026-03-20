"""
excel_store.py — Unico layer di persistenza (sostituisce PostgreSQL).

Architettura:
  • Il file master_cv.xlsx risiede su SharePoint (SHAREPOINT_ROOT_FOLDER/EXCEL_FILENAME)
  • Al boot viene scaricato e tenuto in memoria come dict Python strutturato
  • Ogni scrittura aggiorna la struttura in memoria e ri-uploada il file su SharePoint
  • asyncio.Lock serializza le scritture (sicuro per deploy single-container)

Sheet → struttura dati in memoria (STORE):
  Users        → dict[email, UserRow]
  CVProfiles   → dict[email, CVProfileRow]
  Skills       → dict[email, list[SkillRow]]
  Educations   → dict[email, list[EducationRow]]
  Experiences  → dict[email, list[ExperienceRow]]
  Certifications → dict[email, list[CertificationRow]]
  Languages    → dict[email, list[LanguageRow]]
  Documents    → dict[email, list[DocumentRow]]
  REF - BU     → list[BURow]        (reference data)
  REF - CertTags → list[CertTagRow] (reference data)
  REF - Skills → list[RefSkillRow]  (reference data)

Ogni entità ha un campo "id" generato localmente come UUID4 stringa.
La email è sempre la chiave primaria inter-sheet.
"""
from __future__ import annotations

import asyncio
import io
import re
import uuid
from copy import deepcopy
from datetime import date, datetime, timezone
from typing import Any, Optional

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic_settings import BaseSettings


# ═══════════════════════════════════════════════════════════════════════════════
# SETTINGS
# ═══════════════════════════════════════════════════════════════════════════════

class Settings(BaseSettings):
    secret_key: str = "dev-secret-key-change-in-production"
    access_token_expire_minutes: int = 720
    auth_provider: str = "fake"
    backdoor_password: str = "changeme-set-in-env"
    cors_origins: str = "http://localhost:8082"
    ai_service_url: str = "http://ai-services:8000"
    upload_dir: str = "/app/uploads"
    max_upload_size_mb: int = 10

    entra_tenant_id: Optional[str] = None
    entra_client_id: Optional[str] = None
    entra_client_secret: Optional[str] = None
    entra_audience: Optional[str] = None
    entra_redirect_uri: str = "http://localhost:8082/auth/callback"

    sharepoint_site_url: str = ""
    sharepoint_drive_name: str = "Documenti"
    sharepoint_root_folder: str = "STAFF_DATA_AND_DOCUMENTS"

    excel_filename: str = "master_cv.xlsx"
    excel_backup_filename: str = "master_cv_backup.xlsx"

    @property
    def entra_enabled(self) -> bool:
        return bool(self.entra_tenant_id and self.entra_client_id and self.entra_client_secret)

    @property
    def sharepoint_enabled(self) -> bool:
        return bool(self.entra_enabled and self.sharepoint_site_url)

    model_config = {"env_file": ".env", "extra": "ignore"}


settings = Settings()

# ═══════════════════════════════════════════════════════════════════════════════
# COSTANTI SHEET
# ═══════════════════════════════════════════════════════════════════════════════

SHEET_STAFF        = "Staff"      # foglio unificato Users + CVProfiles
SHEET_USERS        = "Users"      # legacy (lettura backward-compat)
SHEET_CVPROFILES   = "CVProfiles" # legacy (lettura backward-compat)
SHEET_SKILLS       = "Skills"
SHEET_EDUCATIONS   = "Educations"
SHEET_EXPERIENCES  = "Experiences"
SHEET_CERTS        = "Certifications"
SHEET_LANGUAGES    = "Languages"
SHEET_DOCUMENTS    = "Documents"
SHEET_REF_BU       = "REF - BU"
SHEET_REF_CERTTAGS = "REF - CertTags"
SHEET_REF_SKILLS   = "REF - Skills"

# Campi utente (account) nel foglio Staff
_STAFF_USER_FIELDS  = ["id","email","full_name","username","role",
                        "is_active","bu_mashfrog","mashfrog_office","hire_date",
                        "created_at","updated_at"]
# Campi profilo CV nel foglio Staff (cv_updated_at = updated_at di cv_profiles)
_STAFF_CV_FIELDS    = ["title","summary","phone","linkedin_url","birth_date",
                        "birth_place","residence_city","first_employment_date",
                        "availability_status","cv_updated_at"]

HEADERS = {
    SHEET_STAFF:        _STAFF_USER_FIELDS + _STAFF_CV_FIELDS,
    SHEET_USERS:        ["id","email","full_name","username","role",
                         "is_active","bu_mashfrog","mashfrog_office","hire_date",
                         "created_at","updated_at"],
    SHEET_CVPROFILES:   ["email","title","summary","phone","linkedin_url","birth_date",
                         "birth_place","residence_city","first_employment_date",
                         "availability_status","updated_at"],
    SHEET_SKILLS:       ["id","email","skill_name","category","rating","notes"],
    SHEET_EDUCATIONS:   ["id","email","institution","degree_level","field_of_study",
                         "graduation_year","graduation_date","grade","notes"],
    SHEET_EXPERIENCES:  ["id","email","company_name","client_name","role","start_date",
                         "end_date","is_current","project_description","activities",
                         "sort_order"],
    SHEET_CERTS:        ["id","email","name","issuing_org","cert_code","version","year",
                         "expiry_date","has_formal_cert","doc_attachment_type","doc_url",
                         "credly_badge_id","badge_image_url","uploaded_file_path","tags","notes"],
    SHEET_LANGUAGES:    ["id","email","language_name","level"],
    SHEET_DOCUMENTS:    ["id","email","doc_type","original_filename","sharepoint_path",
                         "sharepoint_url","upload_date","ai_updated","tags"],
    SHEET_REF_BU:       ["bu_name","description"],
    SHEET_REF_CERTTAGS: ["tag","area","description"],
    SHEET_REF_SKILLS:   ["skill_name","category","notes","usage_count"],
}

MIME_EXCEL = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ═══════════════════════════════════════════════════════════════════════════════
# IN-MEMORY STORE
# ═══════════════════════════════════════════════════════════════════════════════

# Struttura principale: dizionari per accesso O(1) per email
STORE: dict[str, Any] = {
    "users":         {},   # email → dict
    "cv_profiles":   {},   # email → dict
    "skills":        {},   # email → list[dict]
    "educations":    {},   # email → list[dict]
    "experiences":   {},   # email → list[dict]
    "certifications":{},   # email → list[dict]
    "languages":     {},   # email → list[dict]
    "documents":     {},   # email → list[dict]
    "ref_bu":        [],   # list[dict]
    "ref_certtags":  [],   # list[dict]
    "ref_skills":    [],   # list[dict]
}

_write_lock = asyncio.Lock()
_initialized = False
_last_backup_time: Optional[datetime] = None
_BACKUP_INTERVAL_SECONDS = 2 * 3600   # 2 ore

# Write-Ahead: True se persist() ha fallito e lo STORE RAM è più recente di SharePoint.
# Il loop di retry in main.py tenta di risincronizzare ogni 30s finché _dirty è True.
_dirty: bool = False


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def new_id() -> str:
    return str(uuid.uuid4())


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _fmt(val: Any) -> str:
    """Serializza un valore per cella Excel."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "SI" if val else "NO"
    if isinstance(val, (date, datetime)):
        return val.isoformat()
    if isinstance(val, list):
        return " | ".join(str(v) for v in val)
    return str(val)


def _parse_bool(val: str) -> bool:
    return str(val).upper() in ("SI", "TRUE", "1", "YES")


def _parse_list(val: str) -> list[str]:
    if not val or not str(val).strip():
        return []
    return [v.strip() for v in str(val).split("|") if v.strip()]


def _parse_int(val: Any) -> Optional[int]:
    try:
        return int(val) if val not in (None, "") else None
    except (ValueError, TypeError):
        return None


def _today_iso() -> str:
    return date.today().isoformat()


# ═══════════════════════════════════════════════════════════════════════════════
# SHAREPOINT I/O
# ═══════════════════════════════════════════════════════════════════════════════

_sp_token_cache: dict = {}
_drive_id_cache: str = ""


def _parse_site(url: str) -> tuple[str, str]:
    url = url.rstrip("/")
    m = re.match(r"https?://([^/]+)(/.+)", url)
    if not m:
        raise ValueError(f"URL SharePoint non valido: {url}")
    return m.group(1), m.group(2)


async def _get_token() -> str:
    global _sp_token_cache
    now = datetime.now(timezone.utc).timestamp()
    if _sp_token_cache.get("expires_at", 0) > now + 60:
        return _sp_token_cache["token"]

    url = f"https://login.microsoftonline.com/{settings.entra_tenant_id}/oauth2/v2.0/token"
    async with httpx.AsyncClient(timeout=15) as client:
        resp = await client.post(url, data={
            "grant_type": "client_credentials",
            "client_id": settings.entra_client_id,
            "client_secret": settings.entra_client_secret,
            "scope": "https://graph.microsoft.com/.default",
        })
        resp.raise_for_status()
        data = resp.json()

    _sp_token_cache = {
        "token": data["access_token"],
        "expires_at": now + data.get("expires_in", 3600),
    }
    return _sp_token_cache["token"]


async def _get_drive_id() -> str:
    global _drive_id_cache
    if _drive_id_cache:
        return _drive_id_cache

    token = await _get_token()
    host, site_path = _parse_site(settings.sharepoint_site_url)
    async with httpx.AsyncClient(timeout=15) as client:
        site_resp = await client.get(
            f"https://graph.microsoft.com/v1.0/sites/{host}:{site_path}",
            headers={"Authorization": f"Bearer {token}"},
        )
        site_resp.raise_for_status()
        site_id = site_resp.json()["id"]

        resp = await client.get(
            f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives",
            headers={"Authorization": f"Bearer {token}"},
        )
        resp.raise_for_status()
        drives = resp.json().get("value", [])

    for d in drives:
        if d.get("name") == settings.sharepoint_drive_name:
            _drive_id_cache = d["id"]
            return _drive_id_cache

    available = [d.get("name") for d in drives]
    raise ValueError(f"Drive '{settings.sharepoint_drive_name}' non trovato. Disponibili: {available}")


async def _sp_download(filename: str) -> Optional[bytes]:
    """Scarica un file da SharePoint. Ritorna None se non esiste (404)."""
    if not settings.sharepoint_enabled:
        return None
    try:
        token = await _get_token()
        drive_id = await _get_drive_id()
        url = (f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
               f"/root:/{settings.sharepoint_root_folder}/{filename}:/content")
        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
            resp = await client.get(url, headers={"Authorization": f"Bearer {token}"})
            if resp.status_code == 404:
                return None
            resp.raise_for_status()
            return resp.content
    except Exception as e:
        print(f"[SP] Download '{filename}' fallito: {e}")
        return None


async def _sp_upload(filename: str, content: bytes, _retries: int = 4) -> Optional[str]:
    """Carica un file su SharePoint. Ritorna il webUrl.
    Ritenta fino a _retries volte in caso di 423 Locked (file aperto da un altro utente).
    """
    if not settings.sharepoint_enabled:
        return None
    token = await _get_token()
    drive_id = await _get_drive_id()
    url = (f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
           f"/root:/{settings.sharepoint_root_folder}/{filename}:/content")
    last_error: Optional[Exception] = None
    for attempt in range(1, _retries + 1):
        try:
            async with httpx.AsyncClient(timeout=60) as client:
                resp = await client.put(
                    url,
                    headers={"Authorization": f"Bearer {token}", "Content-Type": MIME_EXCEL},
                    content=content,
                )
                if resp.status_code == 423:
                    wait = attempt * 3
                    print(f"[SP] Upload '{filename}' — 423 Locked (tentativo {attempt}/{_retries}), riprovo in {wait}s...")
                    await asyncio.sleep(wait)
                    continue
                resp.raise_for_status()
                return resp.json().get("webUrl", "")
        except Exception as e:
            last_error = e
            print(f"[SP] Upload '{filename}' errore (tentativo {attempt}/{_retries}): {e}")
            if attempt < _retries:
                await asyncio.sleep(attempt * 2)
    print(f"[SP] Upload '{filename}' fallito dopo {_retries} tentativi: {last_error}")
    return None


async def _sp_delete(filename: str) -> bool:
    """Elimina un file da SharePoint. Ritorna True se eliminato (o già assente)."""
    if not settings.sharepoint_enabled:
        return False
    try:
        token    = await _get_token()
        drive_id = await _get_drive_id()
        # Prima ottieni l'item ID tramite il path
        url_meta = (f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
                    f"/root:/{settings.sharepoint_root_folder}/{filename}")
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.get(url_meta, headers={"Authorization": f"Bearer {token}"})
            if r.status_code == 404:
                return True   # già non esiste
            r.raise_for_status()
            item_id = r.json()["id"]
            url_del = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}"
            rd = await client.delete(url_del, headers={"Authorization": f"Bearer {token}"})
            return rd.status_code in (204, 404)
    except Exception as e:
        print(f"[SP] Delete '{filename}' fallito: {e}")
        return False


async def _sp_get_modified_date(filename: str) -> Optional[str]:
    """Ritorna la data di ultima modifica del file (YYYY-MM-DD) o None se non esiste."""
    if not settings.sharepoint_enabled:
        return None
    try:
        token = await _get_token()
        drive_id = await _get_drive_id()
        url = (f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
               f"/root:/{settings.sharepoint_root_folder}/{filename}")
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(url, headers={"Authorization": f"Bearer {token}"})
            if resp.status_code == 404:
                return None
            resp.raise_for_status()
            last_mod = resp.json().get("lastModifiedDateTime", "")
            return last_mod[:10] if last_mod else None
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════════════════════
# WORKBOOK → STORE / STORE → WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

def _wb_to_store(wb: openpyxl.Workbook) -> None:
    """Legge il workbook openpyxl e popola STORE."""
    global STORE
    STORE = {k: ({} if isinstance(v, dict) else []) for k, v in STORE.items()}

    def read_sheet(sheet_name: str) -> list[dict]:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            result.append({headers[i]: (str(row[i]).strip() if row[i] is not None else "")
                           for i in range(len(headers))})
        return result

    # Staff (foglio unificato Users + CVProfiles) — con fallback ai fogli legacy
    if SHEET_STAFF in wb.sheetnames:
        for r in read_sheet(SHEET_STAFF):
            email = r.get("email", "").lower().strip()
            if not email:
                continue
            # Campi account utente
            user_row = {f: r.get(f, "") for f in _STAFF_USER_FIELDS}
            STORE["users"][email] = user_row
            # Campi profilo CV (cv_updated_at → updated_at per compat interna)
            cv_row  = {f: r.get(f, "") for f in ["title","summary","phone","linkedin_url",
                        "birth_date","birth_place","residence_city",
                        "first_employment_date","availability_status"]}
            cv_row["email"]      = email
            cv_row["updated_at"] = r.get("cv_updated_at", "")
            STORE["cv_profiles"][email] = cv_row
    else:
        # Backward-compat: leggi i due fogli separati
        for r in read_sheet(SHEET_USERS):
            email = r.get("email", "").lower().strip()
            if email:
                STORE["users"][email] = r
        for r in read_sheet(SHEET_CVPROFILES):
            email = r.get("email", "").lower().strip()
            if email:
                STORE["cv_profiles"][email] = r

    # Liste per email
    for sheet, store_key in [
        (SHEET_SKILLS,       "skills"),
        (SHEET_EDUCATIONS,   "educations"),
        (SHEET_EXPERIENCES,  "experiences"),
        (SHEET_CERTS,        "certifications"),
        (SHEET_LANGUAGES,    "languages"),
        (SHEET_DOCUMENTS,    "documents"),
    ]:
        for r in read_sheet(sheet):
            email = r.get("email", "")
            if email:
                STORE[store_key].setdefault(email, []).append(r)

    # Reference data
    STORE["ref_bu"]       = read_sheet(SHEET_REF_BU)
    STORE["ref_certtags"] = read_sheet(SHEET_REF_CERTTAGS)
    STORE["ref_skills"]   = read_sheet(SHEET_REF_SKILLS)


def _store_to_wb() -> openpyxl.Workbook:
    """Serializza STORE in un workbook openpyxl."""
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")
    REF_FILL    = PatternFill("solid", fgColor="E2EFDA")

    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]

    def add_sheet(name: str, headers: list[str], rows: list[list], is_ref: bool = False) -> None:
        ws = wb.create_sheet(title=name)
        ws.freeze_panes = "A2"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = REF_FILL if is_ref else HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, row in enumerate(rows, start=2):
            ws.append([_fmt(v) for v in row])
            if not is_ref and i % 2 == 0:
                for cell in ws[i]:
                    cell.fill = ALT_FILL
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # Staff — foglio unificato Users + CVProfiles
    staff_headers = HEADERS[SHEET_STAFF]
    staff_rows = []
    for email, u in STORE["users"].items():
        cv = STORE["cv_profiles"].get(email, {})
        row = [u.get(f, "") for f in _STAFF_USER_FIELDS] + \
              [cv.get(f, "") if f != "cv_updated_at" else cv.get("updated_at", "")
               for f in _STAFF_CV_FIELDS]
        staff_rows.append(row)
    add_sheet(SHEET_STAFF, staff_headers, staff_rows)

    # Liste per email
    for sheet, store_key in [
        (SHEET_SKILLS,       "skills"),
        (SHEET_EDUCATIONS,   "educations"),
        (SHEET_EXPERIENCES,  "experiences"),
        (SHEET_CERTS,        "certifications"),
        (SHEET_LANGUAGES,    "languages"),
        (SHEET_DOCUMENTS,    "documents"),
    ]:
        headers = HEADERS[sheet]
        rows = [
            [item.get(h, "") for h in headers]
            for items in STORE[store_key].values()
            for item in items
        ]
        add_sheet(sheet, headers, rows)

    # REF sheets
    for sheet, store_key, is_ref in [
        (SHEET_REF_BU,       "ref_bu",       True),
        (SHEET_REF_CERTTAGS, "ref_certtags", True),
        (SHEET_REF_SKILLS,   "ref_skills",   True),
    ]:
        headers = HEADERS[sheet]
        rows = [[r.get(h, "") for h in headers] for r in STORE[store_key]]
        add_sheet(sheet, headers, rows, is_ref=True)

    return wb


def _wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# INIT & PERSIST
# ═══════════════════════════════════════════════════════════════════════════════

async def init_store() -> None:
    """Chiamato al boot: scarica master_cv.xlsx da SharePoint e popola STORE."""
    global _initialized
    if _initialized:
        return

    print(f"[Store] Inizializzazione da SharePoint ({settings.excel_filename})...")
    content = await _sp_download(settings.excel_filename)

    if content:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        _wb_to_store(wb)
        print(f"[Store] Caricati: "
              f"{len(STORE['users'])} utenti, "
              f"{sum(len(v) for v in STORE['skills'].values())} skill, "
              f"{sum(len(v) for v in STORE['experiences'].values())} esperienze")
    else:
        print("[Store] File non trovato su SharePoint — store vuoto (primo avvio).")

    _initialized = True


async def persist() -> bool:
    """Serializza STORE → Excel → upload SharePoint (chiamato dopo ogni write).

    Ritorna True se la sincronizzazione è riuscita (o SharePoint è disabilitato).
    Ritorna False e setta _dirty=True se tutti i retry su SP falliscono.
    Non solleva eccezioni: i dati restano comunque aggiornati in memoria.
    """
    global _dirty
    wb = _store_to_wb()
    content = _wb_to_bytes(wb)

    if not settings.sharepoint_enabled:
        _dirty = False
        return True

    url = await _sp_upload(settings.excel_filename, content)
    if url is not None:
        if _dirty:
            print("[Store] WAL: dati precedentemente non salvati ora sincronizzati su SharePoint.")
        _dirty = False
        return True

    # Upload fallito dopo tutti i retry (es. file locked > 30s)
    _dirty = True
    print("[Store] WAL: persist() fallita — STORE RAM aggiornato, SharePoint non sincronizzato. "
          "Retry automatico ogni 30s.")
    return False


async def retry_persist_if_dirty() -> bool:
    """Ritenta persist() se _dirty=True. Acquisisce il write lock.

    Chiamare solo dall'esterno del lock (es. dal loop di retry in main.py).
    Ritorna True se non c'era nulla da fare o se il retry è riuscito.
    """
    if not _dirty:
        return True
    async with _write_lock:
        if not _dirty:   # ricontrollo dopo aver acquisito il lock
            return True
        print("[Store] WAL retry: tentativo di sincronizzare dati non salvati su SharePoint...")
        return await persist()


async def do_periodic_backup() -> bool:
    """
    Crea un backup ogni _BACKUP_INTERVAL_SECONDS (default 2 ore).
    Copia master_cv.xlsx → master_cv_backup.xlsx su SharePoint.
    """
    global _last_backup_time
    if not settings.sharepoint_enabled:
        return False

    now = datetime.utcnow()
    if _last_backup_time is not None:
        elapsed = (now - _last_backup_time).total_seconds()
        if elapsed < _BACKUP_INTERVAL_SECONDS:
            return False

    print(f"[Backup] Avvio backup periodico ({int(_BACKUP_INTERVAL_SECONDS/3600)}h)...")
    content = await _sp_download(settings.excel_filename)
    if content:
        await _sp_upload(settings.excel_backup_filename, content)
        _last_backup_time = now
        print(f"[Backup] Backup aggiornato: {settings.excel_backup_filename}")
        return True
    return False


# Alias per backward-compat con eventuali chiamate esterne
do_daily_backup = do_periodic_backup


async def _sp_list_folder(folder_path: str) -> list[dict]:
    """
    Elenca i file in una cartella SharePoint.
    folder_path: percorso relativo sotto sharepoint_root_folder
    Ritorna lista di dict con 'name', 'size', 'lastModifiedDateTime'.
    """
    if not settings.sharepoint_enabled:
        return []
    try:
        token = await _get_token()
        drive_id = await _get_drive_id()
        url = (f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
               f"/root:/{settings.sharepoint_root_folder}/{folder_path}:/children")
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.get(
                url,
                headers={"Authorization": f"Bearer {token}"},
            )
            if resp.status_code == 404:
                return []
            resp.raise_for_status()
            items = resp.json().get("value", [])
            return [
                {"name": i["name"], "size": i.get("size", 0),
                 "lastModifiedDateTime": i.get("lastModifiedDateTime", "")}
                for i in items
                if not i.get("folder")   # solo file, non sotto-cartelle
            ]
    except Exception as e:
        print(f"[SP] Listing '{folder_path}' fallito: {e}")
        return []


async def sp_download_file(folder_path: str, filename: str) -> Optional[bytes]:
    """Scarica un singolo file da una sotto-cartella SharePoint."""
    return await _sp_download(f"{folder_path}/{filename}")


# ═══════════════════════════════════════════════════════════════════════════════
# API — USERS
# ═══════════════════════════════════════════════════════════════════════════════

def get_user(email: str) -> Optional[dict]:
    return STORE["users"].get(email)


def get_user_by_username(username: str) -> Optional[dict]:
    for u in STORE["users"].values():
        if u.get("username") == username:
            return u
    return None


def list_users() -> list[dict]:
    return list(STORE["users"].values())


async def create_user(data: dict) -> dict:
    async with _write_lock:
        email = data["email"]
        if email in STORE["users"]:
            raise ValueError(f"Utente {email} già esistente")
        now = now_iso()
        row = {
            "id": new_id(),
            "email": email,
            "full_name": data.get("full_name", ""),
            "username": data.get("username", email.split("@")[0]),
            "role": data.get("role", "USER"),
            "is_active": "SI",
            "bu_mashfrog": data.get("bu_mashfrog", ""),
            "mashfrog_office": data.get("mashfrog_office", ""),
            "hire_date": data.get("hire_date", ""),
            "created_at": now,
            "updated_at": now,
        }
        STORE["users"][email] = row
        # Crea profilo CV vuoto
        STORE["cv_profiles"][email] = {
            "email": email, "title": "", "summary": "", "phone": "",
            "linkedin_url": "", "birth_date": "", "birth_place": "",
            "residence_city": "", "first_employment_date": "",
            "availability_status": "IN_STAFF", "updated_at": now,
        }
        await persist()
        return row


async def update_user(email: str, data: dict) -> dict:
    async with _write_lock:
        row = STORE["users"].get(email)
        if not row:
            raise KeyError(f"Utente {email} non trovato")
        for k, v in data.items():
            if k in row:
                row[k] = _fmt(v) if not isinstance(v, str) else v
        row["updated_at"] = now_iso()
        await persist()
        return row


async def delete_user(email: str) -> None:
    async with _write_lock:
        if email not in STORE["users"]:
            raise KeyError(f"Utente {email} non trovato")
        STORE["users"].pop(email, None)
        STORE["cv_profiles"].pop(email, None)
        for key in ("skills","educations","experiences","certifications","languages","documents"):
            STORE[key].pop(email, None)
        await persist()


# ═══════════════════════════════════════════════════════════════════════════════
# API — CV PROFILES
# ═══════════════════════════════════════════════════════════════════════════════

def get_cv_profile(email: str) -> Optional[dict]:
    return STORE["cv_profiles"].get(email)


async def update_cv_profile(email: str, data: dict) -> dict:
    async with _write_lock:
        row = STORE["cv_profiles"].setdefault(email, {"email": email})
        for k, v in data.items():
            row[k] = _fmt(v) if not isinstance(v, str) else v
        row["updated_at"] = now_iso()
        # Aggiorna updated_at anche sullo user
        if email in STORE["users"]:
            STORE["users"][email]["updated_at"] = row["updated_at"]
        await persist()
        return row


# ═══════════════════════════════════════════════════════════════════════════════
# API — SKILL
# ═══════════════════════════════════════════════════════════════════════════════

def get_skills(email: str) -> list[dict]:
    return STORE["skills"].get(email, [])


async def add_skill(email: str, data: dict) -> dict:
    async with _write_lock:
        row = {"id": new_id(), "email": email,
               "skill_name": data.get("skill_name",""),
               "category": data.get("category","HARD"),
               "rating": str(data.get("rating","")) ,
               "notes": data.get("notes","")}
        STORE["skills"].setdefault(email, []).append(row)
        await persist()
        return row


async def update_skill(email: str, skill_id: str, data: dict) -> dict:
    async with _write_lock:
        items = STORE["skills"].get(email, [])
        row = next((r for r in items if r["id"] == skill_id), None)
        if not row:
            raise KeyError(f"Skill {skill_id} non trovato")
        row.update({k: (str(v) if v is not None else "") for k, v in data.items() if k in row})
        await persist()
        return row


async def delete_skill(email: str, skill_id: str) -> None:
    async with _write_lock:
        items = STORE["skills"].get(email, [])
        STORE["skills"][email] = [r for r in items if r["id"] != skill_id]
        await persist()


# ═══════════════════════════════════════════════════════════════════════════════
# API — EDUCATION
# ═══════════════════════════════════════════════════════════════════════════════

def get_educations(email: str) -> list[dict]:
    return STORE["educations"].get(email, [])


async def add_education(email: str, data: dict) -> dict:
    async with _write_lock:
        row = {"id": new_id(), "email": email,
               "institution":    data.get("institution","") or "",
               "degree_level":   data.get("degree_level","") or "",
               "field_of_study": data.get("field_of_study","") or "",
               "graduation_year":str(data.get("graduation_year","") or ""),
               "graduation_date":str(data.get("graduation_date","") or ""),
               "grade":          data.get("grade","") or "",
               "notes":          data.get("notes","") or ""}
        STORE["educations"].setdefault(email, []).append(row)
        await persist()
        return row


async def update_education(email: str, edu_id: str, data: dict) -> dict:
    async with _write_lock:
        items = STORE["educations"].get(email, [])
        row = next((r for r in items if r["id"] == edu_id), None)
        if not row:
            raise KeyError(f"Education {edu_id} non trovata")
        row.update({k: (str(v) if v is not None else "") for k, v in data.items() if k in row})
        await persist()
        return row


async def delete_education(email: str, edu_id: str) -> None:
    async with _write_lock:
        items = STORE["educations"].get(email, [])
        STORE["educations"][email] = [r for r in items if r["id"] != edu_id]
        await persist()


# ═══════════════════════════════════════════════════════════════════════════════
# API — EXPERIENCES
# ═══════════════════════════════════════════════════════════════════════════════

def get_experiences(email: str) -> list[dict]:
    items = STORE["experiences"].get(email, [])
    # Ordina: is_current prima, poi end_date DESC, poi start_date DESC
    def sort_key(r):
        is_cur = r.get("is_current","NO").upper() in ("SI","TRUE","1")
        end = r.get("end_date","") or "0000"
        start = r.get("start_date","") or "0000"
        return (not is_cur, end == "", end, start)
    return sorted(items, key=sort_key, reverse=False)


async def add_experience(email: str, data: dict) -> dict:
    async with _write_lock:
        row = {"id": new_id(), "email": email,
               "company_name":       data.get("company_name","") or "",
               "client_name":        data.get("client_name","") or "",
               "role":               data.get("role","") or "",
               "start_date":         str(data.get("start_date","") or ""),
               "end_date":           str(data.get("end_date","") or ""),
               "is_current":         _fmt(data.get("is_current", False)),
               "project_description":data.get("project_description","") or "",
               "activities":         data.get("activities","") or "",
               "sort_order":         str(data.get("sort_order","0") or "0")}
        STORE["experiences"].setdefault(email, []).append(row)
        await persist()
        return row


async def update_experience(email: str, exp_id: str, data: dict) -> dict:
    async with _write_lock:
        items = STORE["experiences"].get(email, [])
        row = next((r for r in items if r["id"] == exp_id), None)
        if not row:
            raise KeyError(f"Experience {exp_id} non trovata")
        for k, v in data.items():
            if k in row:
                row[k] = _fmt(v) if not isinstance(v, str) else v
        await persist()
        return row


async def delete_experience(email: str, exp_id: str) -> None:
    async with _write_lock:
        items = STORE["experiences"].get(email, [])
        STORE["experiences"][email] = [r for r in items if r["id"] != exp_id]
        await persist()


# ═══════════════════════════════════════════════════════════════════════════════
# API — CERTIFICATIONS
# ═══════════════════════════════════════════════════════════════════════════════

def get_certifications(email: str) -> list[dict]:
    return STORE["certifications"].get(email, [])


async def add_certification(email: str, data: dict) -> dict:
    async with _write_lock:
        tags = data.get("tags", [])
        row = {"id": new_id(), "email": email,
               "name":        data.get("name","") or "",
               "issuing_org": data.get("issuing_org","") or "",
               "cert_code":   data.get("cert_code","") or "",
               "version":     data.get("version","") or "",
               "year":        str(data.get("year","") or ""),
               "expiry_date": str(data.get("expiry_date","") or ""),
               "has_formal_cert": _fmt(data.get("has_formal_cert", True)),
               "doc_attachment_type": data.get("doc_attachment_type","NONE"),
               "doc_url": data.get("doc_url",""),
               "credly_badge_id": data.get("credly_badge_id",""),
               "badge_image_url": data.get("badge_image_url",""),
               "uploaded_file_path": data.get("uploaded_file_path",""),
               "tags": " | ".join(tags) if isinstance(tags, list) else str(tags or ""),
               "notes": data.get("notes","")}
        STORE["certifications"].setdefault(email, []).append(row)
        await persist()
        return row


async def update_certification(email: str, cert_id: str, data: dict) -> dict:
    async with _write_lock:
        items = STORE["certifications"].get(email, [])
        row = next((r for r in items if r["id"] == cert_id), None)
        if not row:
            raise KeyError(f"Certification {cert_id} non trovata")
        _allowed = set(SHEET_COLUMNS[SHEET_CERTS])
        for k, v in data.items():
            if k in row or k in _allowed:
                if k == "tags" and isinstance(v, list):
                    row[k] = " | ".join(v)
                else:
                    row[k] = _fmt(v) if not isinstance(v, str) else v
        await persist()
        return row


async def delete_certification(email: str, cert_id: str) -> None:
    async with _write_lock:
        items = STORE["certifications"].get(email, [])
        STORE["certifications"][email] = [r for r in items if r["id"] != cert_id]
        await persist()


# ═══════════════════════════════════════════════════════════════════════════════
# API — LANGUAGES
# ═══════════════════════════════════════════════════════════════════════════════

def get_languages(email: str) -> list[dict]:
    return STORE["languages"].get(email, [])


async def add_language(email: str, data: dict) -> dict:
    async with _write_lock:
        row = {"id": new_id(), "email": email,
               "language_name": data.get("language_name",""),
               "level": data.get("level","")}
        STORE["languages"].setdefault(email, []).append(row)
        await persist()
        return row


async def update_language(email: str, lang_id: str, data: dict) -> dict:
    async with _write_lock:
        items = STORE["languages"].get(email, [])
        row = next((r for r in items if r["id"] == lang_id), None)
        if not row:
            raise KeyError(f"Language {lang_id} non trovata")
        row.update({k: str(v or "") for k, v in data.items() if k in row})
        await persist()
        return row


async def delete_language(email: str, lang_id: str) -> None:
    async with _write_lock:
        items = STORE["languages"].get(email, [])
        STORE["languages"][email] = [r for r in items if r["id"] != lang_id]
        await persist()


# ═══════════════════════════════════════════════════════════════════════════════
# API — DOCUMENTS
# ═══════════════════════════════════════════════════════════════════════════════

def get_documents(email: str) -> list[dict]:
    return STORE["documents"].get(email, [])


async def add_document(email: str, data: dict) -> dict:
    async with _write_lock:
        tags = data.get("tags", [])
        row = {"id": new_id(), "email": email,
               "doc_type": data.get("doc_type","UPLOAD"),
               "original_filename": data.get("original_filename",""),
               "sharepoint_path": data.get("sharepoint_path",""),
               "sharepoint_url": data.get("sharepoint_url",""),
               "upload_date": data.get("upload_date", now_iso()),
               "ai_updated": _fmt(data.get("ai_updated", False)),
               "tags": " | ".join(tags) if isinstance(tags, list) else str(tags or "")}
        STORE["documents"].setdefault(email, []).append(row)
        await persist()
        return row


async def update_document(email: str, doc_id: str, data: dict) -> dict:
    async with _write_lock:
        items = STORE["documents"].get(email, [])
        row = next((r for r in items if r["id"] == doc_id), None)
        if not row:
            raise KeyError(f"Document {doc_id} non trovato")
        for k, v in data.items():
            if k in row:
                if k == "tags" and isinstance(v, list):
                    row[k] = " | ".join(v)
                else:
                    row[k] = _fmt(v) if not isinstance(v, str) else v
        await persist()
        return row


async def delete_document(email: str, doc_id: str) -> bool:
    sp_path = None
    async with _write_lock:
        items = STORE["documents"].get(email, [])
        doc = next((r for r in items if r["id"] == doc_id), None)
        if doc is None:
            return False
        sp_path = doc.get("sharepoint_path", "")
        STORE["documents"][email] = [r for r in items if r["id"] != doc_id]
        await persist()
    # Cancella da SP fuori dal lock (non bloccante)
    if sp_path:
        await _sp_delete(sp_path)
    return True


# ═══════════════════════════════════════════════════════════════════════════════
# API — REFERENCE DATA
# ═══════════════════════════════════════════════════════════════════════════════

def get_ref_bu() -> list[dict]:
    return STORE["ref_bu"]


def get_ref_certtags() -> list[dict]:
    return STORE["ref_certtags"]


def get_ref_skills() -> list[dict]:
    return STORE["ref_skills"]


# ═══════════════════════════════════════════════════════════════════════════════
# UTILITY — COMPLETENESS SCORE
# ═══════════════════════════════════════════════════════════════════════════════

def compute_completeness(email: str) -> float:
    profile = STORE["cv_profiles"].get(email, {})
    user    = STORE["users"].get(email, {})
    score   = 0.0
    checks  = [
        bool(profile.get("title")),
        bool(profile.get("summary")),
        bool(profile.get("phone")),
        bool(profile.get("linkedin_url")),
        bool(profile.get("birth_date")),
        bool(profile.get("residence_city")),
        bool(profile.get("first_employment_date")),
        bool(user.get("bu_mashfrog")),
        bool(user.get("mashfrog_office")),
        len(STORE["skills"].get(email, [])) > 0,
        len(STORE["educations"].get(email, [])) > 0,
        len(STORE["experiences"].get(email, [])) > 0,
        len(STORE["certifications"].get(email, [])) > 0,
        len(STORE["languages"].get(email, [])) > 0,
    ]
    return round(sum(checks) / len(checks), 3)   # 0.0 – 1.0


# ═══════════════════════════════════════════════════════════════════════════════
# UTILITY — SKILL AUTOCOMPLETE
# ═══════════════════════════════════════════════════════════════════════════════

def suggest_skills(query: str, limit: int = 10) -> list[str]:
    q = query.lower().strip()
    seen: dict[str, int] = {}
    for items in STORE["skills"].values():
        for s in items:
            name = s.get("skill_name","")
            if q in name.lower():
                seen[name] = seen.get(name, 0) + 1
    return [k for k, _ in sorted(seen.items(), key=lambda x: -x[1])][:limit]


def suggest_cert_codes(query: str, email: str, limit: int = 10) -> list[dict]:
    """Suggerisce codici certificazione basandosi sulle cert già nel sistema (Jaccard)."""
    from difflib import SequenceMatcher

    q_tokens = set(re.sub(r"[^\w]", " ", query.lower()).split())
    stop_words = {"the","and","or","of","in","for","to","a","an","with","on","at"}
    q_tokens -= stop_words

    candidates: dict[str, dict] = {}
    for certs in STORE["certifications"].values():
        for c in certs:
            code = c.get("cert_code","")
            name = c.get("name","")
            if not code or not name:
                continue
            c_tokens = set(re.sub(r"[^\w]", " ", name.lower()).split()) - stop_words
            if not c_tokens or not q_tokens:
                continue
            jaccard = len(q_tokens & c_tokens) / len(q_tokens | c_tokens)
            seq = SequenceMatcher(None, query.lower(), name.lower()).ratio()
            score = max(jaccard, seq)
            if score >= 0.55 and code not in candidates:
                candidates[code] = {"cert_code": code, "name": name,
                                    "issuing_org": c.get("issuing_org",""), "score": score}

    return sorted(candidates.values(), key=lambda x: -x["score"])[:limit]
