"""
Carica i dati dal file People_Analytics nel database.
Eseguito automaticamente all'avvio se il file esiste in /app/data/.
"""
import re
import logging
from datetime import date
from pathlib import Path
from typing import Optional, List, Tuple

import openpyxl
from sqlalchemy.orm import Session

from app.models import (
    User, CV, Education, Language, CVRole, CVSkill, Reference,
    Certification, SkillCategory, DegreeLevel, LanguageLevel,
    DocAttachmentType, UserRole, AvailabilityStatus,
)
logger = logging.getLogger(__name__)

EXCEL_PATH = Path("/app/data/people_analytics.xlsx")
EMAIL_DOMAIN = "mashfrog.com"
_PLACEHOLDER_HASH = "SEED_PLACEHOLDER"  # sovrascritto da _sync_all_passwords in seed.py


# ── Date / year helpers ───────────────────────────────────────────────────────

def parse_date(val) -> Optional[date]:
    if val is None:
        return None
    # Already a date/datetime object (openpyxl)
    if hasattr(val, "date"):
        return val.date()
    if isinstance(val, date):
        return val
    # Integer year
    if isinstance(val, (int, float)):
        y = int(val)
        return date(y, 1, 1) if 1900 <= y <= 2100 else None
    s = str(val).strip()
    if not s or s.lower() in ("none", ""):
        return None
    # "YYYY-MM-DD HH:MM:SS" or "YYYY-MM-DD"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # "DD/MM/YYYY"
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", s)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    # "YYYY" alone
    m = re.match(r"^(\d{4})$", s)
    if m:
        return date(int(m.group(1)), 1, 1)
    # "A/A YYYY/YY" academic year
    m = re.match(r"A/A\s*(\d{4})", s, re.IGNORECASE)
    if m:
        return date(int(m.group(1)), 9, 1)
    return None


def parse_year(val) -> Optional[int]:
    d = parse_date(val)
    return d.year if d else None


# ── Skill parsing ─────────────────────────────────────────────────────────────

def parse_skills(text: str, category: SkillCategory) -> List[Tuple[str, Optional[int]]]:
    """Returns list of (skill_name, rating) where rating is 1-5 or None."""
    if not text or str(text).strip().lower() in ("none", ""):
        return []
    results: List[Tuple[str, Optional[int]]] = []
    for raw in str(text).splitlines():
        line = raw.strip().lstrip("•-–\t").strip()
        if not line:
            continue
        name, rating = _parse_skill_line(line)
        if name and len(name) >= 2:
            results.append((name, rating))
    return results


def _parse_skill_line(line: str) -> Tuple[Optional[str], Optional[int]]:
    # "N - Name" (rating first)
    m = re.match(r"^(\d)\s*[-–]\s*(.+)$", line)
    if m:
        return m.group(2).strip(), int(m.group(1))
    # "Name - N" (rating last, only one digit at end)
    m = re.match(r"^(.+?)\s*[-–]\s*(\d)\s*$", line)
    if m:
        return m.group(1).strip(), int(m.group(2))
    # "Name: N" or "Name:N"
    m = re.match(r"^(.+?)\s*:\s*(\d)\s*$", line)
    if m:
        return m.group(1).strip(), int(m.group(2))
    # "Name (N)"
    m = re.match(r"^(.+?)\s*\((\d)\)\s*$", line)
    if m:
        return m.group(1).strip(), int(m.group(2))
    # No rating — just the name (skip long free-text)
    if len(line) < 80:
        return line, None
    return None, None


# ── Language parsing ──────────────────────────────────────────────────────────

_LANG_LEVEL_MAP = {
    "madre": LanguageLevel.MADRELINGUA,
    "madrelingua": LanguageLevel.MADRELINGUA,
    "madre lingua": LanguageLevel.MADRELINGUA,
    "lingua madre": LanguageLevel.MADRELINGUA,
    "a1": LanguageLevel.A1, "a2": LanguageLevel.A2,
    "b1": LanguageLevel.B1, "b2": LanguageLevel.B2,
    "c1": LanguageLevel.C1, "c2": LanguageLevel.C2,
}


def parse_languages(text: str) -> List[Tuple[str, Optional[LanguageLevel]]]:
    if not text or str(text).strip().lower() in ("none", ""):
        return []
    results: List[Tuple[str, Optional[LanguageLevel]]] = []
    for raw in str(text).splitlines():
        line = raw.strip().lstrip("-–\t").strip()
        if not line:
            continue
        parts = re.split(r"\s*[-–]\s*", line, maxsplit=1)
        lang_name = parts[0].strip().title()
        level = None
        if len(parts) > 1:
            lvl_str = parts[1].strip().lower()
            # Direct lookup
            level = _LANG_LEVEL_MAP.get(lvl_str)
            if not level:
                m = re.search(r"\b([abc][12])\b", lvl_str, re.IGNORECASE)
                if m:
                    level = _LANG_LEVEL_MAP.get(m.group(1).lower())
                elif "madre" in lvl_str:
                    level = LanguageLevel.MADRELINGUA
        if lang_name and len(lang_name) >= 2:
            results.append((lang_name, level))
    return results


# ── Certification parsing ─────────────────────────────────────────────────────

def parse_certifications(text: str) -> List[dict]:
    if not text or str(text).strip().lower() in ("none", ""):
        return []
    results = []
    for raw in str(text).splitlines():
        line = raw.strip().lstrip("-–\t").strip()
        if not line or line in ("[..]", ".", "-"):
            continue
        # "YYYY: Description"
        m = re.match(r"^(\d{4})\s*:\s*(.+)$", line)
        if m:
            results.append({"year": int(m.group(1)), "name": m.group(2).strip()})
            continue
        # "YYYY. Description" or "YYYY - Description" or "YYYY – Description"
        m = re.match(r"^(\d{4})\s*[.\-–]\s*(.+)$", line)
        if m:
            results.append({"year": int(m.group(1)), "name": m.group(2).strip()})
            continue
        # Plain description (no year)
        if len(line) > 3:
            results.append({"year": None, "name": line})
    return results


# ── Experience / reference parsing ────────────────────────────────────────────

def parse_experiences(text: str) -> List[dict]:
    """
    Parses multi-line experience text into structured records.
    Returns list of {start_year, end_year, is_current, description}.
    """
    if not text or str(text).strip().lower() in ("none", ""):
        return []
    records = []
    for raw in str(text).splitlines():
        line = raw.strip().lstrip("-–\t").strip()
        if not line:
            continue
        # "YYYY - YYYY - Description" or "YYYY - Attuale - Description"
        m = re.match(
            r"^(\d{4})\s*[-–/]\s*(Attuale|\d{4}|oggi|ad oggi)\s*[-–]\s*(.+)$",
            line, re.IGNORECASE,
        )
        if m:
            start = int(m.group(1))
            end_str = m.group(2).strip().lower()
            is_curr = end_str in ("attuale", "oggi", "ad oggi")
            end_y = None if is_curr else (int(end_str) if end_str.isdigit() else None)
            records.append({
                "start_year": start, "end_year": end_y,
                "is_current": is_curr,
                "description": m.group(3).strip(),
            })
            continue
        # "MM/YYYY–oggi ..." or "MM/YYYY-Attuale ..."
        m = re.match(
            r"^(\d{2})/(\d{4})\s*[-–]\s*(oggi|Attuale)(.+)?$",
            line, re.IGNORECASE,
        )
        if m:
            records.append({
                "start_year": int(m.group(2)), "end_year": None,
                "is_current": True,
                "description": (m.group(4) or "").strip(),
            })
            continue
        # Continuation of previous record or standalone
        if records:
            records[-1]["description"] = (
                records[-1].get("description", "") + " " + line
            ).strip()
        else:
            records.append({
                "start_year": None, "end_year": None,
                "is_current": False,
                "description": line,
            })
    return records


# ── Utility ───────────────────────────────────────────────────────────────────

def username_to_fullname(username: str) -> str:
    parts = username.split(".")
    return " ".join(p.capitalize() for p in parts)


def _cell(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s.lower() not in ("none", "") else None


# ── Main seeding function ─────────────────────────────────────────────────────

def seed_from_excel(db: Session, excel_path: Path = EXCEL_PATH) -> None:
    if not excel_path.exists():
        logger.warning(
            "Excel seed file not found at %s — Excel seeding skipped. "
            "Copy people_analytics.xlsx to backend/data/ to enable.",
            excel_path,
        )
        return

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb.active

    # Locate the header row (row containing "STAFF" in column B)
    data_start_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[1] and str(row[1]).strip().upper() == "STAFF":
            data_start_row = i + 1
            break

    if data_start_row is None:
        logger.error("Header row not found in Excel — aborting Excel seed")
        return

    logger.info("Excel seed: starting from row %d", data_start_row)
    inserted = 0
    skipped = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < data_start_row:
            continue

        username = _cell(row[1])
        if not username:
            continue

        email = f"{username}@{EMAIL_DOMAIN}"

        # Skip if already exists
        if db.query(User).filter(User.email == email).first():
            logger.debug("Skip existing user: %s", email)
            skipped += 1
            continue

        # ── User ──────────────────────────────────────────────────────────────
        user = User(
            email=email,
            username=username,
            full_name=username_to_fullname(username),
            hashed_password=_PLACEHOLDER_HASH,  # password vera impostata da _sync_all_passwords
            role=UserRole.USER,
            is_active=True,
            bu_mashfrog=_cell(row[0]),
            mashfrog_office=_cell(row[5]),
        )
        db.add(user)
        db.flush()

        # ── CV ─────────────────────────────────────────────────────────────────
        cv = CV(
            user_id=user.id,
            birth_date=parse_date(row[2]),
            birth_place=_cell(row[3]),
            residence_city=_cell(row[4]),
            first_employment_date=parse_date(row[17]),
            title=_cell(row[15]),
            availability_status=AvailabilityStatus.DISPONIBILE,
        )
        db.add(cv)
        db.flush()

        # ── Education — Diploma ────────────────────────────────────────────────
        diploma = _cell(row[6])
        if diploma:
            db.add(Education(
                cv_id=cv.id,
                degree_level=DegreeLevel.DIPLOMA,
                institution=diploma,
                graduation_year=parse_year(row[7]),
                graduation_date=parse_date(row[7]),
            ))

        # ── Education — Laurea Triennale ──────────────────────────────────────
        triennale = _cell(row[8])
        if triennale:
            db.add(Education(
                cv_id=cv.id,
                degree_level=DegreeLevel.TRIENNALE,
                institution=triennale,
                graduation_year=parse_year(row[9]),
                graduation_date=parse_date(row[9]),
            ))

        # ── Education — Laurea Magistrale ─────────────────────────────────────
        magistrale = _cell(row[10])
        if magistrale:
            db.add(Education(
                cv_id=cv.id,
                degree_level=DegreeLevel.MAGISTRALE,
                institution=magistrale,
                graduation_year=parse_year(row[11]),
                graduation_date=parse_date(row[11]),
            ))

        # ── Education — Altri Master / Corsi (multi-line) ─────────────────────
        altri = _cell(row[12])
        if altri:
            for line in altri.splitlines():
                line = line.strip()
                if not line:
                    continue
                m = re.match(r"^(\d{4})[.\-\s]+(.+)$", line)
                if m:
                    db.add(Education(
                        cv_id=cv.id,
                        degree_level=DegreeLevel.CORSO,
                        institution=m.group(2).strip()[:255],
                        graduation_year=int(m.group(1)),
                    ))
                else:
                    db.add(Education(
                        cv_id=cv.id,
                        degree_level=DegreeLevel.CORSO,
                        institution=line[:255],
                    ))

        # ── Languages ─────────────────────────────────────────────────────────
        seen_langs = set()
        for lang_name, level in parse_languages(str(row[13] or "")):
            if lang_name.lower() in seen_langs:
                continue
            seen_langs.add(lang_name.lower())
            db.add(Language(cv_id=cv.id, language_name=lang_name, level=level))

        # ── CVRole — ruolo attuale ─────────────────────────────────────────────
        ruolo = _cell(row[15])
        if ruolo:
            db.add(CVRole(
                cv_id=cv.id,
                title=ruolo[:255],
                start_date=parse_date(row[16]),
                is_current=True,
                company="Mashfrog",
            ))

        # ── Hard Skills ────────────────────────────────────────────────────────
        seen_skills: set = set()
        for skill_name, rating in parse_skills(str(row[19] or ""), SkillCategory.HARD):
            key = (skill_name.lower()[:50], "HARD")
            if key in seen_skills:
                continue
            seen_skills.add(key)
            if rating is not None and not (1 <= rating <= 5):
                rating = None
            db.add(CVSkill(
                cv_id=cv.id,
                skill_name=skill_name[:255],
                category=SkillCategory.HARD,
                rating=rating,
            ))

        # ── Soft Skills ────────────────────────────────────────────────────────
        for skill_name, rating in parse_skills(str(row[20] or ""), SkillCategory.SOFT):
            key = (skill_name.lower()[:50], "SOFT")
            if key in seen_skills:
                continue
            seen_skills.add(key)
            if rating is not None and not (1 <= rating <= 5):
                rating = None
            db.add(CVSkill(
                cv_id=cv.id,
                skill_name=skill_name[:255],
                category=SkillCategory.SOFT,
                rating=rating,
            ))

        # ── Certifications ────────────────────────────────────────────────────
        for cert in parse_certifications(str(row[14] or "")):
            db.add(Certification(
                cv_id=cv.id,
                year=cert.get("year"),
                name=cert["name"][:500],
                doc_attachment_type=DocAttachmentType.NONE,
                has_formal_cert=True,
            ))

        # ── References (esperienze) ────────────────────────────────────────────
        sintesi = str(row[18] or "").strip()
        if sintesi and sintesi.lower() != "none":
            exp_records = parse_experiences(sintesi)
            if exp_records:
                for idx, exp in enumerate(exp_records[:15]):
                    start_d = date(exp["start_year"], 1, 1) if exp.get("start_year") else None
                    end_d = date(exp["end_year"], 12, 31) if exp.get("end_year") else None
                    db.add(Reference(
                        cv_id=cv.id,
                        start_date=start_d,
                        end_date=end_d,
                        is_current=exp.get("is_current", False),
                        company_name="Mashfrog",
                        project_description=exp.get("description", "")[:2000],
                        sort_order=idx,
                    ))
            else:
                # Fallback: store entire text as single reference
                db.add(Reference(
                    cv_id=cv.id,
                    is_current=True,
                    company_name="Mashfrog",
                    project_description=sintesi[:2000],
                    sort_order=0,
                ))

        inserted += 1

    db.commit()
    logger.info(
        "Excel seed completed: %d inserted, %d skipped (already existed)",
        inserted, skipped,
    )
