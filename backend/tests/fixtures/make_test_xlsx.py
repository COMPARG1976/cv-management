"""
Script per generare tests/fixtures/test_master_cv.xlsx.

Eseguire da dentro il container:
    docker exec cv_mgmt_backend python tests/fixtures/make_test_xlsx.py

Oppure direttamente:
    python backend/tests/fixtures/make_test_xlsx.py
"""
import io, sys
from pathlib import Path

# Assicura import da backend/
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import openpyxl
from app.excel_store import (
    HEADERS,
    SHEET_STAFF, SHEET_SKILLS, SHEET_EDUCATIONS, SHEET_EXPERIENCES,
    SHEET_CERTS, SHEET_LANGUAGES, SHEET_DOCUMENTS,
    SHEET_REF_BU, SHEET_REF_CERTTAGS, SHEET_REF_SKILLS,
)

OUTPUT = Path(__file__).parent / "test_master_cv.xlsx"

# ─── Dati di test ─────────────────────────────────────────────────────────────

USERS = [
    {
        "id": "u-0001",
        "email": "test.user@example.com",
        "full_name": "Test User",
        "username": "test.user",
        "role": "USER",
        "is_active": "SI",
        "bu_mashfrog": "Development",
        "mashfrog_office": "Roma",
        "hire_date": "2022-01-10",
        "created_at": "2022-01-10T10:00:00",
        "updated_at": "2022-01-10T10:00:00",
        # CV fields
        "title": "Software Engineer",
        "summary": "Sviluppatore Python con 5 anni di esperienza.",
        "phone": "+39 333 1234567",
        "linkedin_url": "https://linkedin.com/in/testuser",
        "birth_date": "1990-05-15",
        "birth_place": "Roma",
        "residence_city": "Roma",
        "first_employment_date": "2018-06-01",
        "availability_status": "IN_STAFF",
        "cv_updated_at": "2024-01-01T12:00:00",
    },
    {
        "id": "u-0002",
        "email": "test.admin@example.com",
        "full_name": "Test Admin",
        "username": "test.admin",
        "role": "ADMIN",
        "is_active": "SI",
        "bu_mashfrog": "Management",
        "mashfrog_office": "Roma",
        "hire_date": "2020-01-01",
        "created_at": "2020-01-01T08:00:00",
        "updated_at": "2020-01-01T08:00:00",
        # CV fields
        "title": "IT Manager",
        "summary": "",
        "phone": "",
        "linkedin_url": "",
        "birth_date": "",
        "birth_place": "",
        "residence_city": "Milano",
        "first_employment_date": "",
        "availability_status": "IN_STAFF",
        "cv_updated_at": "2024-01-01T12:00:00",
    },
]

SKILLS = [
    {"id": "sk-001", "email": "test.user@example.com", "skill_name": "Python",      "category": "HARD", "rating": 5, "notes": ""},
    {"id": "sk-002", "email": "test.user@example.com", "skill_name": "FastAPI",     "category": "HARD", "rating": 4, "notes": ""},
    {"id": "sk-003", "email": "test.user@example.com", "skill_name": "Teamwork",    "category": "SOFT", "rating": 4, "notes": ""},
    {"id": "sk-004", "email": "test.admin@example.com","skill_name": "Leadership",  "category": "SOFT", "rating": 5, "notes": ""},
]

EDUCATIONS = [
    {
        "id": "ed-001", "email": "test.user@example.com",
        "institution": "Università La Sapienza", "degree_level": "LAUREA_MAGISTRALE",
        "field_of_study": "Informatica", "graduation_year": 2015,
        "graduation_date": "", "grade": "110/110", "notes": "",
    },
]

EXPERIENCES = [
    {
        "id": "ex-001", "email": "test.user@example.com",
        "company_name": "Mashfrog", "client_name": "Cliente ABC",
        "role": "Backend Developer", "start_date": "2022-01-01",
        "end_date": "", "is_current": "SI",
        "project_description": "Sviluppo API REST per gestione CV.",
        "activities": "Python, FastAPI, PostgreSQL",
        "sort_order": 0,
    },
    {
        "id": "ex-002", "email": "test.user@example.com",
        "company_name": "Acme Corp", "client_name": "",
        "role": "Junior Developer", "start_date": "2018-06-01",
        "end_date": "2021-12-01", "is_current": "NO",
        "project_description": "Sviluppo applicazioni web.",
        "activities": "Java, Spring Boot",
        "sort_order": 1,
    },
]

CERTIFICATIONS = [
    {
        "id": "ce-001", "email": "test.user@example.com",
        "name": "AWS Certified Solutions Architect",
        "issuing_org": "Amazon Web Services",
        "cert_code": "SAA-C03", "version": "2023",
        "year": 2023, "expiry_date": "2026-01-01",
        "has_formal_cert": "SI",
        "doc_attachment_type": "URL",
        "doc_url": "https://www.credly.com/badges/test",
        "credly_badge_id": "", "uploaded_file_path": "",
        "tags": "Cloud;AWS", "notes": "",
    },
]

LANGUAGES = [
    {"id": "la-001", "email": "test.user@example.com", "language_name": "Italiano", "level": "MADRELINGUA"},
    {"id": "la-002", "email": "test.user@example.com", "language_name": "Inglese",  "level": "B2"},
]

REF_BU = [
    {"bu_name": "Development",  "description": ""},
    {"bu_name": "Management",   "description": ""},
    {"bu_name": "Data & AI",    "description": ""},
    {"bu_name": "Cloud",        "description": ""},
]

REF_CERTTAGS = [
    {"tag": "Cloud",  "area": "Infrastructure", "description": ""},
    {"tag": "AWS",    "area": "Cloud",           "description": ""},
    {"tag": "Python", "area": "Development",     "description": ""},
]

REF_SKILLS = [
    {"skill_name": "Python",    "category": "HARD", "notes": "", "usage_count": 10},
    {"skill_name": "FastAPI",   "category": "HARD", "notes": "", "usage_count": 5},
    {"skill_name": "Java",      "category": "HARD", "notes": "", "usage_count": 3},
    {"skill_name": "Leadership","category": "SOFT", "notes": "", "usage_count": 8},
    {"skill_name": "Teamwork",  "category": "SOFT", "notes": "", "usage_count": 6},
]


# ─── Builder ──────────────────────────────────────────────────────────────────

def _write_sheet(wb: openpyxl.Workbook, sheet_name: str, headers: list, rows: list[dict]):
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def build_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # rimuove sheet di default

    _write_sheet(wb, SHEET_STAFF, HEADERS[SHEET_STAFF], USERS)
    _write_sheet(wb, SHEET_SKILLS, HEADERS[SHEET_SKILLS], SKILLS)
    _write_sheet(wb, SHEET_EDUCATIONS, HEADERS[SHEET_EDUCATIONS], EDUCATIONS)
    _write_sheet(wb, SHEET_EXPERIENCES, HEADERS[SHEET_EXPERIENCES], EXPERIENCES)
    _write_sheet(wb, SHEET_CERTS, HEADERS[SHEET_CERTS], CERTIFICATIONS)
    _write_sheet(wb, SHEET_LANGUAGES, HEADERS[SHEET_LANGUAGES], LANGUAGES)
    _write_sheet(wb, SHEET_DOCUMENTS, HEADERS[SHEET_DOCUMENTS], [])
    _write_sheet(wb, SHEET_REF_BU, HEADERS[SHEET_REF_BU], REF_BU)
    _write_sheet(wb, SHEET_REF_CERTTAGS, HEADERS[SHEET_REF_CERTTAGS], REF_CERTTAGS)
    _write_sheet(wb, SHEET_REF_SKILLS, HEADERS[SHEET_REF_SKILLS], REF_SKILLS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    data = build_xlsx()
    OUTPUT.write_bytes(data)
    print(f"Generato: {OUTPUT}  ({len(data):,} bytes)")
    # Verifica rapida
    wb = openpyxl.load_workbook(io.BytesIO(data))
    for sh in wb.sheetnames:
        print(f"  {sh}: {wb[sh].max_row} righe")
