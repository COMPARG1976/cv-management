import os
"""
import_to_master_excel.py
=========================
Importa il file di export PostgreSQL (cv_management_data_20260318_1510.xlsx)
da SharePoint e lo converte nel formato master_cv.xlsx (Excel-only backend).

Operazioni:
  1. Scarica il file sorgente da SharePoint (stessa cartella del master_cv.xlsx)
  2. Legge tutti i fogli e converte il formato:
     - Users: rimuove hashed_password, converte ID interi ? UUID
     - Tutti gli altri sheet: converte ID interi ? UUID
     - Tags: gi? pipe-separated nell'export
  3. Salva come master_cv.xlsx e carica su SharePoint

Uso:
    cd C:/20.PROGETTI_CLAUDE_CODE/40.CV_MANAGEMENT
    python scripts/import_to_master_excel.py [--source nome_file.xlsx] [--dry-run]

    --source    nome del file sorgente su SharePoint (default: cv_management_data_20260318_1510.xlsx)
    --dry-run   esegue la conversione ma NON carica su SharePoint (salva in /tmp/master_cv_preview.xlsx)
"""

import argparse
import asyncio
import io
import sys
import uuid
from pathlib import Path

# Path setup per importare app.*
BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR / "backend"))

import openpyxl

# Importa il modulo store (legge .env automaticamente)
import app.excel_store as store
from app.excel_store import (
    settings, STORE, HEADERS,
    SHEET_USERS, SHEET_CVPROFILES, SHEET_SKILLS, SHEET_EDUCATIONS,
    SHEET_EXPERIENCES, SHEET_CERTS, SHEET_LANGUAGES, SHEET_DOCUMENTS,
    SHEET_REF_BU, SHEET_REF_CERTTAGS, SHEET_REF_SKILLS,
    _sp_download, _sp_upload, _store_to_wb, _wb_to_bytes,
    _wb_to_store, new_id, now_iso,
)


# ?? Mappa fogli vecchio formato ? chiavi STORE ??????????????????????????????

SHEET_STORE_MAP = [
    (SHEET_SKILLS,       "skills"),
    (SHEET_EDUCATIONS,   "educations"),
    (SHEET_EXPERIENCES,  "experiences"),
    (SHEET_CERTS,        "certifications"),
    (SHEET_LANGUAGES,    "languages"),
    (SHEET_DOCUMENTS,    "documents"),
]

# Fogli di riferimento (REF - xxx) che potrebbero avere nomi diversi
REF_FALLBACKS = {
    SHEET_REF_BU:       ["[REF] BU",       "REF_BU",    "BU"],
    SHEET_REF_CERTTAGS: ["[REF] CertTags", "CertTags"],
    SHEET_REF_SKILLS:   ["[REF] Skills",   "RefSkills", "Skills_REF"],
}


def _read_sheet(wb: openpyxl.Workbook, name: str,
                fallbacks: list[str] | None = None) -> list[dict]:
    """Legge un foglio del workbook per nome (prova fallback se non trovato)."""
    candidates = [name] + (fallbacks or [])
    ws = None
    for c in candidates:
        if c in wb.sheetnames:
            ws = wb[c]
            break
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        result.append({
            headers[i]: (str(row[i]).strip() if row[i] is not None else "")
            for i in range(len(headers))
        })
    return result


def _ensure_uuid(val: str) -> str:
    """Se val ? un intero o stringa vuota, genera un nuovo UUID; altrimenti lo usa as-is."""
    if not val:
        return new_id()
    try:
        int(val)          # ? un intero DB ? sostituisci con UUID
        return new_id()
    except ValueError:
        pass
    try:
        uuid.UUID(val)    # gi? UUID valido ? ok
        return val
    except ValueError:
        return new_id()   # formato sconosciuto ? nuovo UUID


def convert_old_to_store(wb: openpyxl.Workbook) -> None:
    """
    Legge il workbook vecchio formato e popola STORE con i dati convertiti.
    Differenze rispetto a _wb_to_store:
      - Users: salta hashed_password, converte IDs
      - Tutti: converte IDs interi ? UUID
    """
    global STORE
    # Reset STORE
    for key in STORE:
        if isinstance(STORE[key], dict):
            STORE[key] = {}
        else:
            STORE[key] = []

    # ?? Users ??????????????????????????????????????????????????????????????????
    new_headers = HEADERS[SHEET_USERS]   # senza hashed_password
    for r in _read_sheet(wb, SHEET_USERS):
        email = r.get("email", "").lower().strip()
        if not email:
            continue
        row = {
            "id":              _ensure_uuid(r.get("id", "")),
            "email":           email,
            "full_name":       r.get("full_name", ""),
            "username":        r.get("username", email.split("@")[0]),
            "role":            r.get("role", "USER"),
            "is_active":       r.get("is_active", "SI"),
            "bu_mashfrog":     r.get("bu_mashfrog", ""),
            "mashfrog_office": r.get("mashfrog_office", ""),
            "hire_date":       r.get("hire_date", ""),
            "created_at":      r.get("created_at", "") or now_iso(),
            "updated_at":      r.get("updated_at", "") or now_iso(),
        }
        STORE["users"][email] = row

    # ?? CVProfiles ?????????????????????????????????????????????????????????????
    for r in _read_sheet(wb, SHEET_CVPROFILES):
        email = r.get("email", "").lower().strip()
        if not email:
            continue
        STORE["cv_profiles"][email] = {k: r.get(k, "") for k in HEADERS[SHEET_CVPROFILES]}
        STORE["cv_profiles"][email]["email"] = email

    # ?? Skills ?????????????????????????????????????????????????????????????????
    for r in _read_sheet(wb, SHEET_SKILLS):
        email = r.get("email", "").lower().strip()
        if not email:
            continue
        row = {k: r.get(k, "") for k in HEADERS[SHEET_SKILLS]}
        row["id"] = _ensure_uuid(r.get("id", ""))
        row["email"] = email
        STORE["skills"].setdefault(email, []).append(row)

    # ?? Educations ?????????????????????????????????????????????????????????????
    for r in _read_sheet(wb, SHEET_EDUCATIONS):
        email = r.get("email", "").lower().strip()
        if not email:
            continue
        row = {k: r.get(k, "") for k in HEADERS[SHEET_EDUCATIONS]}
        row["id"] = _ensure_uuid(r.get("id", ""))
        row["email"] = email
        STORE["educations"].setdefault(email, []).append(row)

    # ?? Experiences ????????????????????????????????????????????????????????????
    for r in _read_sheet(wb, SHEET_EXPERIENCES):
        email = r.get("email", "").lower().strip()
        if not email:
            continue
        row = {k: r.get(k, "") for k in HEADERS[SHEET_EXPERIENCES]}
        row["id"] = _ensure_uuid(r.get("id", ""))
        row["email"] = email
        STORE["experiences"].setdefault(email, []).append(row)

    # ?? Certifications ?????????????????????????????????????????????????????????
    for r in _read_sheet(wb, SHEET_CERTS):
        email = r.get("email", "").lower().strip()
        if not email:
            continue
        row = {k: r.get(k, "") for k in HEADERS[SHEET_CERTS]}
        row["id"] = _ensure_uuid(r.get("id", ""))
        row["email"] = email
        STORE["certifications"].setdefault(email, []).append(row)

    # ?? Languages ??????????????????????????????????????????????????????????????
    for r in _read_sheet(wb, SHEET_LANGUAGES):
        email = r.get("email", "").lower().strip()
        if not email:
            continue
        row = {k: r.get(k, "") for k in HEADERS[SHEET_LANGUAGES]}
        row["id"] = _ensure_uuid(r.get("id", ""))
        row["email"] = email
        STORE["languages"].setdefault(email, []).append(row)

    # ?? Documents ??????????????????????????????????????????????????????????????
    for r in _read_sheet(wb, SHEET_DOCUMENTS):
        email = r.get("email", "").lower().strip()
        if not email:
            continue
        row = {k: r.get(k, "") for k in HEADERS[SHEET_DOCUMENTS]}
        row["id"] = _ensure_uuid(r.get("id", ""))
        row["email"] = email
        STORE["documents"].setdefault(email, []).append(row)

    # ?? Profili CV mancanti: crea vuoto per ogni utente senza profilo ??????????
    for email in list(STORE["users"]):
        if email not in STORE["cv_profiles"]:
            STORE["cv_profiles"][email] = {
                "email": email, "title": "", "summary": "", "phone": "",
                "linkedin_url": "", "birth_date": "", "birth_place": "",
                "residence_city": "", "first_employment_date": "",
                "availability_status": "IN_STAFF", "updated_at": now_iso(),
            }

    # ?? REF sheets ?????????????????????????????????????????????????????????????
    STORE["ref_bu"]       = _read_sheet(wb, SHEET_REF_BU,
                                        REF_FALLBACKS[SHEET_REF_BU])
    STORE["ref_certtags"] = _read_sheet(wb, SHEET_REF_CERTTAGS,
                                        REF_FALLBACKS[SHEET_REF_CERTTAGS])
    STORE["ref_skills"]   = _read_sheet(wb, SHEET_REF_SKILLS,
                                        REF_FALLBACKS[SHEET_REF_SKILLS])


def print_summary() -> None:
    print("\n  ? Riepilogo dati convertiti:")
    print(f"     Utenti:        {len(STORE['users'])}")
    print(f"     Profili CV:    {len(STORE['cv_profiles'])}")
    print(f"     Skill:         {sum(len(v) for v in STORE['skills'].values())}")
    print(f"     Formazione:    {sum(len(v) for v in STORE['educations'].values())}")
    print(f"     Esperienze:    {sum(len(v) for v in STORE['experiences'].values())}")
    print(f"     Certificaz.:   {sum(len(v) for v in STORE['certifications'].values())}")
    print(f"     Lingue:        {sum(len(v) for v in STORE['languages'].values())}")
    print(f"     Documenti:     {sum(len(v) for v in STORE['documents'].values())}")
    print(f"     REF BU:        {len(STORE['ref_bu'])}")
    print(f"     REF CertTags:  {len(STORE['ref_certtags'])}")
    print(f"     REF Skills:    {len(STORE['ref_skills'])}")


async def main(source_filename: str, dry_run: bool) -> None:
    print(f"\n{'='*60}")
    print(f"  Import Excel -> master_cv.xlsx")
    print(f"  Sorgente:   {source_filename}")
    print(f"  Modalit?:   {'DRY-RUN (no upload)' if dry_run else 'LIVE (carica su SharePoint)'}")
    print(f"  SharePoint: {settings.sharepoint_root_folder}")
    print(f"{'='*60}\n")

    # ?? 1. Scarica file sorgente ???????????????????????????????????????????????
    print(f"  [1/4] Download '{source_filename}' da SharePoint...")
    content = await _sp_download(source_filename)
    if not content:
        print(f"\n  ERRORE File '{source_filename}' non trovato su SharePoint.")
        print("     Verifica che il file esista nella cartella:")
        print(f"     {settings.sharepoint_root_folder}/")
        sys.exit(1)
    print(f"       ? Scaricato ({len(content):,} bytes)")

    # ?? 2. Carica workbook sorgente ????????????????????????????????????????????
    print(f"  [2/4] Lettura workbook sorgente...")
    wb_src = openpyxl.load_workbook(io.BytesIO(content))
    print(f"       ? Fogli trovati: {wb_src.sheetnames}")

    # ?? 3. Converti in STORE ???????????????????????????????????????????????????
    print(f"  [3/4] Conversione dati...")
    convert_old_to_store(wb_src)
    print_summary()

    # ?? 4. Serializza in nuovo formato ????????????????????????????????????????
    print(f"\n  [4/4] Generazione master_cv.xlsx...")
    wb_new = _store_to_wb()
    new_content = _wb_to_bytes(wb_new)
    print(f"       ? Generato ({len(new_content):,} bytes)")

    if dry_run:
        # Salva in locale per verifica
        local_path = os.path.join(tempfile.gettempdir(), "master_cv_preview.xlsx")
        with open(local_path, "wb") as f:
            f.write(new_content)
        print(f"\n  OK DRY-RUN completato. File salvato in: {local_path}")
        print("     Per caricare su SharePoint, riesegui senza --dry-run")
    else:
        # Carica su SharePoint come master_cv.xlsx
        print(f"       Caricamento su SharePoint come '{settings.excel_filename}'...")
        url = await _sp_upload(settings.excel_filename, new_content)
        if url:
            print(f"  OK Caricato con successo!")
            print(f"     URL: {url}")
        else:
            print(f"  ERRORE Upload fallito ? controlla le credenziali SharePoint in .env")
            sys.exit(1)

    print()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Importa old Excel ? master_cv.xlsx")
    parser.add_argument(
        "--source",
        default="cv_management_data_20260318_1510.xlsx",
        help="Nome del file sorgente su SharePoint (default: cv_management_data_20260318_1510.xlsx)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Non carica su SharePoint, salva solo in /tmp/master_cv_preview.xlsx",
    )
    args = parser.parse_args()

    if not settings.sharepoint_enabled:
        print("\n? SharePoint non configurato in .env")
        print("   Assicurati che ENTRA_TENANT_ID, ENTRA_CLIENT_ID, ENTRA_CLIENT_SECRET")
        print("   e SHAREPOINT_SITE_URL siano valorizzati.")
        sys.exit(1)

    asyncio.run(main(args.source, args.dry_run))
