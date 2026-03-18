"""
export_db_to_excel.py
=====================
Esporta tutti i dati dal DB PostgreSQL in un file Excel strutturato
e lo carica su SharePoint (STAFF_DATA_AND_DOCUMENTS_TEST).

Uso (dentro il container backend oppure con DATABASE_URL settato):
    python scripts/export_db_to_excel.py

Il file viene salvato localmente in /tmp/cv_management_data_YYYYMMDD_HHMM.xlsx
e poi caricato su SharePoint.
"""

import os
import sys
import asyncio
import re
from datetime import datetime
from pathlib import Path

# ── Path setup ──────────────────────────────────────────────────────────────
# Permette di importare app.* sia da dentro il container sia da fuori
BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR / "backend"))

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── DB imports ───────────────────────────────────────────────────────────────
from app.database import SessionLocal
from app.models import (
    User, CV, CVSkill, Education, Reference, Certification,
    Language
)

# ── SharePoint import (opzionale — skippa se non configurato) ────────────────
try:
    from app.sharepoint import _get_token, _get_drive_id, settings as sp_settings
    SHAREPOINT_AVAILABLE = True
except Exception as e:
    print(f"[WARN] SharePoint non disponibile: {e}")
    SHAREPOINT_AVAILABLE = False

# ═══════════════════════════════════════════════════════════════════════════
# STILI
# ═══════════════════════════════════════════════════════════════════════════

HEADER_FILL   = PatternFill("solid", fgColor="1F5C99")   # blu scuro
REF_FILL      = PatternFill("solid", fgColor="2E7D32")   # verde scuro
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
BODY_FONT     = Font(size=10, name="Calibri")
ALT_FILL_1    = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL_2    = PatternFill("solid", fgColor="EEF4FB")
REF_ALT_FILL  = PatternFill("solid", fgColor="E8F5E9")
THIN_BORDER   = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header_row(ws, row_idx: int, is_ref: bool = False):
    fill = REF_FILL if is_ref else HEADER_FILL
    for cell in ws[row_idx]:
        cell.font   = HEADER_FONT
        cell.fill   = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_row(ws, row_idx: int, is_ref: bool = False):
    alt = ALT_FILL_2 if row_idx % 2 == 0 else ALT_FILL_1
    if is_ref:
        alt = REF_ALT_FILL if row_idx % 2 == 0 else ALT_FILL_1
    for cell in ws[row_idx]:
        cell.font      = BODY_FONT
        cell.fill      = alt
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border    = THIN_BORDER


def auto_width(ws, max_width: int = 60):
    """Imposta larghezza colonne in base al contenuto (con limite)."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def add_sheet(wb: openpyxl.Workbook, name: str, headers: list[str],
              rows: list[list], is_ref: bool = False) -> openpyxl.worksheet.worksheet.Worksheet:
    ws = wb.create_sheet(title=name)
    ws.row_dimensions[1].height = 28

    # Header
    ws.append(headers)
    style_header_row(ws, 1, is_ref=is_ref)

    # Dati
    for i, row in enumerate(rows, start=2):
        ws.append(row)
        style_data_row(ws, i, is_ref=is_ref)

    ws.freeze_panes = "A2"
    auto_width(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def fmt_date(d) -> str:
    if d is None:
        return ""
    try:
        return d.strftime("%Y-%m-%d")
    except Exception:
        return str(d)


def fmt_year(y) -> str:
    return str(y) if y else ""


def str_or_empty(v) -> str:
    return str(v) if v is not None else ""


# ═══════════════════════════════════════════════════════════════════════════
# EXPORT PRINCIPALE
# ═══════════════════════════════════════════════════════════════════════════

def build_workbook(db) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    # Rimuovi foglio default
    del wb[wb.sheetnames[0]]

    # ── 1. Users ─────────────────────────────────────────────────────────
    users = db.query(User).order_by(User.email).all()
    user_rows = []
    for u in users:
        user_rows.append([
            u.email,
            u.full_name or "",
            u.username or "",
            u.bu_mashfrog or "",
            u.mashfrog_office or "",
            fmt_date(u.hire_date_mashfrog),
            u.role.value if u.role else "",
            "SI" if u.is_active else "NO",
            fmt_date(u.created_at),
        ])
    add_sheet(wb, "Users",
        ["email", "full_name", "username", "bu_mashfrog", "mashfrog_office",
         "hire_date", "role", "is_active", "created_at"],
        user_rows)
    print(f"  Users: {len(user_rows)} righe")

    # ── 2. CVProfiles ────────────────────────────────────────────────────
    cvs = db.query(CV).join(User).order_by(User.email).all()
    # Mappa user_id → email
    uid2email = {u.id: u.email for u in users}

    cv_rows = []
    for cv in cvs:
        email = uid2email.get(cv.user_id, "")
        cv_rows.append([
            email,
            cv.title or "",
            cv.summary or "",
            cv.phone or "",
            cv.linkedin_url or "",
            fmt_date(cv.birth_date),
            cv.residence_city or "",
            fmt_date(cv.first_employment_date),
            cv.availability_status.value if cv.availability_status else "",
            fmt_date(cv.updated_at),
        ])
    add_sheet(wb, "CVProfiles",
        ["email", "title", "summary", "phone", "linkedin_url", "birth_date",
         "residence_city", "first_employment_date", "availability_status", "last_updated"],
        cv_rows)
    print(f"  CVProfiles: {len(cv_rows)} righe")

    # ── 3. Skills ────────────────────────────────────────────────────────
    skills = db.query(CVSkill).join(CV).join(User).order_by(User.email, CVSkill.skill_name).all()
    # Mappa cv_id → email (tramite uid2email e cv.user_id)
    cvid2email = {cv.id: uid2email.get(cv.user_id, "") for cv in cvs}

    skill_rows = []
    for s in skills:
        email = cvid2email.get(s.cv_id, "")
        skill_rows.append([
            email,
            s.skill_name or "",
            s.category.value if s.category else "",
            s.rating or "",
            s.notes or "",
        ])
    add_sheet(wb, "Skills",
        ["email", "skill_name", "category", "rating_1_5", "notes"],
        skill_rows)
    print(f"  Skills: {len(skill_rows)} righe")

    # ── 4. Educations ────────────────────────────────────────────────────
    educations = db.query(Education).join(CV).join(User).order_by(User.email, Education.graduation_year.desc().nullslast()).all()
    edu_rows = []
    for e in educations:
        email = cvid2email.get(e.cv_id, "")
        edu_rows.append([
            email,
            e.institution or "",
            e.degree_level.value if e.degree_level else "",
            e.field_of_study or "",
            fmt_year(e.graduation_year),
            fmt_date(e.graduation_date),
            e.grade or "",
            e.notes or "",
        ])
    add_sheet(wb, "Educations",
        ["email", "institution", "degree_level", "field_of_study",
         "graduation_year", "graduation_date", "grade", "notes"],
        edu_rows)
    print(f"  Educations: {len(edu_rows)} righe")

    # ── 5. Experiences (modello: Reference) ──────────────────────────────
    experiences = (
        db.query(Reference).join(CV).join(User)
        .order_by(User.email, Reference.end_date.desc().nullslast(), Reference.start_date.desc())
        .all()
    )
    exp_rows = []
    for ex in experiences:
        email = cvid2email.get(ex.cv_id, "")
        exp_rows.append([
            email,
            ex.company_name or "",
            ex.client_name or "",
            ex.role or "",
            fmt_date(ex.start_date),
            fmt_date(ex.end_date),
            "SI" if ex.is_current else "NO",
            ex.project_description or "",
            ex.activities or "",
        ])
    add_sheet(wb, "Experiences",
        ["email", "company_name", "client_name", "role", "start_date", "end_date",
         "is_current", "project_description", "activities"],
        exp_rows)
    print(f"  Experiences: {len(exp_rows)} righe")

    # ── 6. Certifications ────────────────────────────────────────────────
    certifications = db.query(Certification).join(CV).join(User).order_by(User.email, Certification.year.desc().nullslast()).all()
    cert_rows = []
    for c in certifications:
        email = cvid2email.get(c.cv_id, "")
        cert_rows.append([
            email,
            c.name or "",
            c.issuing_org or "",
            c.cert_code or "",
            c.version or "",
            fmt_year(c.year),
            fmt_date(c.expiry_date),
            "SI" if c.has_formal_cert else "NO",
            c.doc_attachment_type.value if c.doc_attachment_type else "",
            c.doc_url or "",
            c.credly_badge_id or "",
            c.uploaded_file_path or "",
            ", ".join(c.tags) if c.tags else "",
            c.notes or "",
        ])
    add_sheet(wb, "Certifications",
        ["email", "name", "issuing_org", "cert_code", "version", "year",
         "expiry_date", "has_formal_cert", "doc_attachment_type", "doc_url",
         "credly_badge_id", "uploaded_file_path", "tags", "notes"],
        cert_rows)
    print(f"  Certifications: {len(cert_rows)} righe")

    # ── 7. Languages ─────────────────────────────────────────────────────
    languages = db.query(Language).join(CV).join(User).order_by(User.email, Language.language_name).all()
    lang_rows = []
    for l in languages:
        email = cvid2email.get(l.cv_id, "")
        lang_rows.append([
            email,
            l.language_name or "",
            l.level.value if l.level else "",
        ])
    add_sheet(wb, "Languages",
        ["email", "language_name", "level"],
        lang_rows)
    print(f"  Languages: {len(lang_rows)} righe")

    # ── 8. Documents ─────────────────────────────────────────────────────
    from app.models import CVDocument
    documents = (
        db.query(CVDocument).join(CV).join(User)
        .order_by(User.email, CVDocument.uploaded_at.desc())
        .all()
    )
    doc_rows = []
    for d in documents:
        email = cvid2email.get(d.cv_id, "")
        tags_str = " | ".join(d.tags) if isinstance(d.tags, list) else (d.tags or "")
        doc_rows.append([
            email,
            "UPLOAD",
            d.original_filename or "",
            d.storage_path or "",
            d.sharepoint_url or "",
            fmt_date(d.uploaded_at),
            "SI" if d.ai_updated else "NO",
            tags_str,
        ])
    add_sheet(wb, "Documents",
        ["email", "doc_type", "original_filename", "sharepoint_path",
         "sharepoint_url", "upload_date", "ai_updated", "tags"],
        doc_rows)
    print(f"  Documents: {len(doc_rows)} righe")

    # ── 9.  REF - BU ─────────────────────────────────────────────────────
    bu_values = sorted({u.bu_mashfrog for u in users if u.bu_mashfrog})
    bu_rows = [[bu, ""] for bu in bu_values]
    # Aggiungi BU standard Mashfrog se non presenti
    STANDARD_BU = [
        "Cloud & Infrastructure", "Cybersecurity", "Data & AI",
        "Digital & Experience", "ERP & Business Applications",
        "Enterprise Architecture", "Project Management Office",
    ]
    existing = {r[0] for r in bu_rows}
    for std in STANDARD_BU:
        if std not in existing:
            bu_rows.append([std, "— standard Mashfrog —"])
    bu_rows.sort(key=lambda x: x[0])
    add_sheet(wb, "REF - BU",
        ["bu_name", "description"],
        bu_rows, is_ref=True)
    print(f"  REF - BU: {len(bu_rows)} righe")

    # ── 9. REF - CertTags ───────────────────────────────────────────────
    CERT_TAGS = [
        ["Microsoft", "Cloud & Infra", "Certificazioni Microsoft Azure / M365 / Power Platform"],
        ["AWS", "Cloud & Infra", "Amazon Web Services"],
        ["GCP", "Cloud & Infra", "Google Cloud Platform"],
        ["OpenText", "ERP & ECM", "OpenText ECM / xECM / Content Suite"],
        ["SAP", "ERP & Business Apps", "SAP (tutti i moduli)"],
        ["Cybersecurity", "Security", "CISSP, CEH, CompTIA Security+, ..."],
        ["Agile / PM", "Project Management", "PMP, PRINCE2, Scrum, SAFe"],
        ["Data & AI", "Data & AI", "Databricks, dbt, ML, AI specializations"],
        ["DevOps", "Engineering", "Kubernetes, Terraform, GitOps"],
        ["Compliance", "Governance", "ISO, GDPR, ITIL"],
    ]
    add_sheet(wb, "REF - CertTags",
        ["tag", "area", "description"],
        CERT_TAGS, is_ref=True)
    print(f"  REF - CertTags: {len(CERT_TAGS)} righe")

    # ── 10. REF - Skills ─────────────────────────────────────────────────
    # Raccoglie skill univoci dal DB + aggiungi colonna categoria
    from sqlalchemy import func
    skill_taxonomy = (
        db.query(CVSkill.skill_name, CVSkill.category,
                 func.count(CVSkill.id).label("usage_count"))
        .group_by(CVSkill.skill_name, CVSkill.category)
        .order_by(CVSkill.category, CVSkill.skill_name)
        .all()
    )
    ref_skill_rows = [
        [s.skill_name, s.category.value if s.category else "", "", str(s.usage_count)]
        for s in skill_taxonomy
    ]
    add_sheet(wb, "REF - Skills",
        ["skill_name", "category", "notes", "usage_count"],
        ref_skill_rows, is_ref=True)
    print(f"  REF - Skills: {len(ref_skill_rows)} righe")

    return wb


# ═══════════════════════════════════════════════════════════════════════════
# UPLOAD SU SHAREPOINT
# ═══════════════════════════════════════════════════════════════════════════

async def upload_to_sharepoint(local_path: str, filename: str) -> str | None:
    """Carica il file su SharePoint e ritorna il webUrl."""
    if not SHAREPOINT_AVAILABLE:
        print("[SKIP] SharePoint non configurato — file solo locale.")
        return None

    import httpx

    try:
        # Leggi file
        with open(local_path, "rb") as f:
            content = f.read()

        # Ottieni token + drive
        token = await _get_token()
        drive_id = await _get_drive_id()

        root_folder = sp_settings.sharepoint_root_folder  # STAFF_DATA_AND_DOCUMENTS_TEST

        # Upload
        url = (
            f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
            f"/root:/{root_folder}/{filename}:/content"
        )
        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.put(
                url,
                headers={
                    "Authorization": f"Bearer {token}",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
                content=content,
            )
            resp.raise_for_status()
            web_url = resp.json().get("webUrl", "")
            print(f"  Upload OK → {web_url}")
            return web_url

    except Exception as e:
        print(f"[ERROR] Upload SharePoint fallito: {e}")
        return None


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

async def main():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename  = f"cv_management_data_{timestamp}.xlsx"
    local_path = f"/tmp/{filename}"

    print(f"\n{'='*60}")
    print(f"  CV Management — Export DB → Excel")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")

    # Leggi DB
    print("Lettura DB PostgreSQL...")
    with SessionLocal() as db:
        wb = build_workbook(db)

    # Salva localmente
    wb.save(local_path)
    print(f"\nFile salvato: {local_path} ({Path(local_path).stat().st_size // 1024} KB)")

    # Upload SharePoint
    print("\nUpload su SharePoint...")
    web_url = await upload_to_sharepoint(local_path, filename)

    print(f"\n{'='*60}")
    if web_url:
        print(f"  Completato!")
        print(f"  SharePoint URL: {web_url}")
    else:
        print(f"  Export locale completato.")
        print(f"  File: {local_path}")
    print(f"{'='*60}\n")

    return local_path, web_url


if __name__ == "__main__":
    asyncio.run(main())
