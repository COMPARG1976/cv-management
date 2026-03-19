"""
_import_certs_analysis.py
=========================
FASE 1 dello script di import certificazioni SAP.

Legge:
  - SAP_CERTIFICAZIONI_2026.xlsx  (Sheet1: Active Certifications, Sheet2: List, Sheet3: Metodologiche)
  - CERTIFICAZIONI_EX_CARTELLA_MADERA/  (133 PDF organizzati per categoria)

Produce:
  - CERT_ANALYSIS_<data>.xlsx  con i seguenti sheet:
      1. CERT_COMPLETO          — tutte le cert da tutte le fonti, una riga per cert×persona
      2. PERSONE_SENZA_EMAIL    — persone in Sheet1 non trovate in Sheet2
      3. PDF_SENZA_MATCH        — PDF senza corrispondenza nell'Excel
      4. CODICI_INCOERENTI      — stessa cert_code con descrizioni diverse
      5. AREE_MASTERDATA        — aree uniche per sheet REF
      6. CLUSTER_MASTERDATA     — cluster unici
      7. AI_VERIFICA            — risultati verifica AI per-PDF (solo con --ai)

USO:
  python _import_certs_analysis.py           # solo analisi testuale
  python _import_certs_analysis.py --ai      # + verifica AI su PDF con match dubbio
  python _import_certs_analysis.py --ai-all  # + verifica AI su TUTTI i PDF abbinati

Il file di output viene scritto nella stessa cartella SharePoint sorgente.
Richiede OPENAI_API_KEY in ambiente o nel file .env del progetto.
"""

import io
import json
import os
import re
import sys
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed

# Forza UTF-8 su stdout (necessario su Windows con cp1252 di default)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

RUN_AI     = "--ai"     in sys.argv   # verifica AI solo su match dubbi
RUN_AI_ALL = "--ai-all" in sys.argv   # verifica AI su tutti i PDF abbinati
if RUN_AI_ALL:
    RUN_AI = True
from collections import defaultdict
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURAZIONE PERCORSI
# ──────────────────────────────────────────────────────────────────────────────

BASE   = r"C:\Users\GiuseppeComparetti\MASHFROG GROUP S.R.L\ENT_SOLUTION_M4P_STAFF - Documenti"
XLS_IN = os.path.join(BASE, "SAP_CERTIFICAZIONI_2026.xlsx")
CERT_DIR = os.path.join(BASE, "CERTIFICAZIONI_EX_CARTELLA_MADERA")
OUT_DIR  = BASE   # output nella stessa cartella SharePoint

TODAY    = date.today().strftime("%Y%m%d")
XLS_OUT  = os.path.join(OUT_DIR, f"CERT_ANALYSIS_{TODAY}.xlsx")

# Cartelle da escludere dalla scansione
SKIP_DIRS = {"RESIGNED", "OLD"}

# Soglie fuzzy matching
THRESH_SURNAME  = 0.82   # match cognome PDF → Sheet2
THRESH_CERTNAME = 0.70   # match nome cert PDF → descrizione Excel


# ──────────────────────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ──────────────────────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Normalizza stringa per confronto: minuscolo, senza accenti, strip."""
    if not s:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower()


def _sim(a: str, b: str) -> float:
    return SequenceMatcher(None, _norm(a), _norm(b)).ratio()


def _parse_year(val) -> str:
    """Normalizza anno/data → stringa anno '2022' o '' se non valida."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        y = int(val)
        return str(y) if 1990 <= y <= 2030 else ""
    if isinstance(val, (datetime, date)):
        return str(val.year)
    s = str(val).strip()
    if s.lower() in ("no", "", "n/a"):
        return ""
    # YYYY-MM-DD ...
    m = re.match(r"(\d{4})", s)
    return m.group(1) if m else s


def _clean(val) -> str:
    """Stringa pulita da None / \xa0."""
    if val is None:
        return ""
    s = str(val).strip().replace("\xa0", "").strip()
    return s if s else ""


def _flag_add(rec: dict, flag: str) -> None:
    """Aggiunge un flag al record senza duplicati."""
    flags = [f.strip() for f in rec["flag"].split("|") if f.strip()]
    if flag not in flags:
        flags.append(flag)
    rec["flag"] = " | ".join(flags)


def _canon_certname(name: str) -> str:
    """Normalizza nome certificazione: dash uniformi, spazi, strip."""
    s = str(name).strip()
    # sostituisce dash unicode con trattino ASCII
    s = re.sub(r"[\u2013\u2014\u2012]", "-", s)
    # compatta spazi multipli
    s = re.sub(r"\s{2,}", " ", s)
    return s.strip()


# ──────────────────────────────────────────────────────────────────────────────
# STEP 1 — CARICA EXCEL SORGENTE
# ──────────────────────────────────────────────────────────────────────────────

print("[1/6] Lettura SAP_CERTIFICAZIONI_2026.xlsx...")
wb_in = openpyxl.load_workbook(XLS_IN, data_only=True)

sheet_names = wb_in.sheetnames
print(f"      Sheet trovati: {sheet_names}")

ws1 = wb_in[sheet_names[0]]   # Active Certifications List
ws2 = wb_in[sheet_names[1]]   # List (con email)
ws3 = wb_in[sheet_names[2]]   # Certificazioni Metodologiche

rows1 = list(ws1.iter_rows(values_only=True))
rows2 = list(ws2.iter_rows(values_only=True))
rows3 = list(ws3.iter_rows(values_only=True))

# Sheet 1: col indices
# 0=Cognome 1=Nome 2=Area 3=CodCert 4=Cluster 5=Descrizione(no header!) 6=Anno 7=Note
S1_COG, S1_NOM, S1_AREA, S1_COD, S1_CLUST, S1_DESC, S1_ANNO, S1_NOTE = 0,1,2,3,4,5,6,7

# Sheet 2: col indices
# 0=STATUS 1=Wave 2=Month 3=AREA 4=Employee 5=S-USER 6=EMAIL
# 7=CertArea 8=CertType 9=Code 10=Cloud 11=SubDone
S2_STA, S2_EMP, S2_EMAIL, S2_TYPE, S2_CODE = 0, 4, 6, 8, 9

# Sheet 3: col indices
# 0=Tipo 1=Descrizione 2=Nominativo 3=DataCons 4=DataScad 5=Numero
S3_TIPO, S3_DESC, S3_NOM, S3_DCONS, S3_DSCAD, S3_NUM = 0,1,2,3,4,5


# ──────────────────────────────────────────────────────────────────────────────
# STEP 2 — BUILD EMAIL MAP DA SHEET 2
# ──────────────────────────────────────────────────────────────────────────────

print("[2/6] Costruzione mappa nome→email da Sheet2...")

# {norm(cognome_nome): email}
email_map: dict[str, str] = {}
# {norm(cognome): [(nome, email)]}  per fuzzy fallback
surname_to_email: dict[str, list] = defaultdict(list)

for row in rows2[1:]:
    emp   = _clean(row[S2_EMP])
    email = _clean(row[S2_EMAIL]).lower()
    if not emp or not email or "@" not in email:
        continue
    key = _norm(emp)
    email_map[key] = email
    # cognome = primo token
    parts = emp.split()
    if parts:
        surname_to_email[_norm(parts[0])].append((emp, email))

print(f"      {len(email_map)} persone con email in Sheet2")

# Anche Sheet3: i nominativi possono dare email aggiuntive
# (li tratteremo per fuzzy match in step 3)

def find_email(cognome: str, nome: str) -> tuple[str, float]:
    """Ritorna (email, confidence). 1.0 = match esatto."""
    key = _norm(f"{cognome} {nome}")
    if key in email_map:
        return email_map[key], 1.0

    # Prova inversione Nome Cognome
    key2 = _norm(f"{nome} {cognome}")
    if key2 in email_map:
        return email_map[key2], 1.0

    # Fuzzy su tutta la mappa
    best_score = 0.0
    best_email = ""
    full = f"{cognome} {nome}"
    for k, em in email_map.items():
        s = _sim(full, k)
        if s > best_score:
            best_score = s
            best_email = em
    if best_score >= 0.85:
        return best_email, round(best_score, 3)

    # Fuzzy solo su cognome
    cog_norm = _norm(cognome)
    cands = surname_to_email.get(cog_norm, [])
    if not cands:
        for sn, lst in surname_to_email.items():
            s = _sim(cognome, sn)
            if s >= 0.88:
                cands = lst
                break
    if len(cands) == 1:
        return cands[0][1], 0.75  # solo cognome match — possibile
    if len(cands) > 1:
        # omonimia: prova nome
        for emp_full, em in cands:
            parts = emp_full.split()
            emp_nome = parts[1] if len(parts) > 1 else ""
            if _sim(nome, emp_nome) >= 0.80:
                return em, 0.80

    return "", 0.0


# ──────────────────────────────────────────────────────────────────────────────
# STEP 3 — SCANSIONE PDF + COGNOME EXTRACTION
# ──────────────────────────────────────────────────────────────────────────────

print("[3/6] Scansione cartella PDF...")

# Lista di PDF: {surname, cert_name_raw, category, filepath_rel}
pdf_records: list[dict] = []

for root, dirs, files in os.walk(CERT_DIR):
    dirs[:] = sorted([d for d in dirs if d not in SKIP_DIRS])
    for fname in sorted(files):
        if not fname.lower().endswith(".pdf"):
            continue
        fpath = os.path.join(root, fname)
        rel   = os.path.relpath(fpath, CERT_DIR)
        category = rel.split(os.sep)[0]  # prima cartella = categoria
        stem  = Path(fname).stem

        # Estrai cognome: tutto dopo l'ultimo _ oppure dopo l'ultimo spazio
        if "_" in stem:
            cert_part, surname = stem.rsplit("_", 1)
        else:
            parts = stem.rsplit(" ", 1)
            cert_part = parts[0] if len(parts) > 1 else stem
            surname   = parts[1] if len(parts) > 1 else ""

        # Pulizia cert name estratto dal filename
        cert_name_raw = cert_part.replace("_", " ").strip()

        pdf_records.append({
            "file":          fname,
            "rel_path":      rel,
            "category":      category,
            "surname_pdf":   surname.upper().strip(),
            "cert_name_raw": cert_name_raw,
        })

print(f"      {len(pdf_records)} PDF trovati")


# ──────────────────────────────────────────────────────────────────────────────
# STEP 4 — BUILD CERT DICT (codice → descrizione canonica + varianti)
# ──────────────────────────────────────────────────────────────────────────────

print("[4/6] Analisi certificazioni e coerenza codici...")

# {cert_code → {desc: count}}
code_to_descs: dict[str, dict] = defaultdict(lambda: defaultdict(int))
# {cert_code → canonical_desc}  (la più frequente)
code_canonical: dict[str, str] = {}
# lista di tutte le cert da Sheet1 + Sheet2
all_certs: list[dict] = []
no_email_people: list[dict] = []

# ── Sheet 1 ──
for row in rows1[1:]:
    cognome = _clean(row[S1_COG])
    nome    = _clean(row[S1_NOM])
    if not cognome:
        continue
    area    = _clean(row[S1_AREA])
    code    = _clean(row[S1_COD])
    cluster = _clean(row[S1_CLUST])
    desc    = _canon_certname(_clean(row[S1_DESC]))
    anno    = _parse_year(row[S1_ANNO] if len(row) > S1_ANNO else None)
    note    = _clean(row[S1_NOTE] if len(row) > S1_NOTE else None)

    email, conf = find_email(cognome, nome)

    flags = []
    if not email:
        flags.append("NO_EMAIL")
    if not code:
        flags.append("NO_CODICE")
    if not anno or anno == "no":
        flags.append("DATA_ANOMALA")
    if code and desc:
        code_to_descs[code][desc] += 1

    rec = {
        "fonte":        "Sheet1",
        "cognome":      cognome,
        "nome":         nome,
        "email":        email,
        "email_conf":   conf,
        "cert_code":    code,
        "cert_name":    desc,
        "area":         area,
        "cluster":      cluster,
        "anno":         anno,
        "note_orig":    note,
        "status":       "",
        "pdf_file":     "",
        "pdf_cat":      "",
        "pdf_conf":     "",
        "flag":         " | ".join(flags),
        "note_rev":     "",   # campo per revisione manuale
    }
    all_certs.append(rec)
    if not email:
        no_email_people.append({"cognome": cognome, "nome": nome})

# ── Sheet 2 ── (aggiunge info email/status; NON duplica se già in Sheet1)
# Costruiamo un set di chiavi Sheet1 per dedup
sheet1_keys = {(_norm(r["cognome"]), _norm(r["cert_code"])) for r in all_certs}

for row in rows2[1:]:
    status = _clean(row[S2_STA])
    emp    = _clean(row[S2_EMP])
    email  = _clean(row[S2_EMAIL]).lower()
    ctype  = _canon_certname(_clean(row[S2_TYPE]))
    code   = _clean(row[S2_CODE])
    if not emp:
        continue

    parts   = emp.split()
    cognome = parts[0] if parts else emp
    nome    = " ".join(parts[1:]) if len(parts) > 1 else ""

    if ((_norm(cognome), _norm(code))) in sheet1_keys:
        # già in Sheet1 — aggiorna solo email/status se mancante
        for rec in all_certs:
            if _norm(rec["cognome"]) == _norm(cognome) and _norm(rec["cert_code"]) == _norm(code):
                if not rec["email"] and email:
                    rec["email"]      = email
                    rec["email_conf"] = 1.0
                    rec["flag"] = rec["flag"].replace("NO_EMAIL", "").strip(" |")
                if not rec["status"]:
                    rec["status"] = status
        if code and ctype:
            code_to_descs[code][ctype] += 1
        continue

    # Cert presente solo in Sheet2
    flags = []
    if status == "RESIGNED":
        flags.append("RESIGNED")
    if not code:
        flags.append("NO_CODICE")
    if code and ctype:
        code_to_descs[code][ctype] += 1

    rec = {
        "fonte":        "Sheet2",
        "cognome":      cognome,
        "nome":         nome,
        "email":        email,
        "email_conf":   1.0,
        "cert_code":    code,
        "cert_name":    ctype,
        "area":         _clean(row[3]),
        "cluster":      "",
        "anno":         "",
        "note_orig":    "",
        "status":       status,
        "pdf_file":     "",
        "pdf_cat":      "",
        "pdf_conf":     "",
        "flag":         " | ".join(flags),
        "note_rev":     "",
    }
    all_certs.append(rec)

# ── Sheet 3 (Metodologiche) ──
for row in rows3[1:]:
    tipo    = _clean(row[S3_TIPO])
    desc    = _canon_certname(_clean(row[S3_DESC]))
    nom     = _clean(row[S3_NOM])
    dcons   = _clean(row[S3_DCONS])
    dscad   = _clean(row[S3_DSCAD])
    num     = _clean(row[S3_NUM] if len(row) > S3_NUM else None)
    if not nom:
        continue

    parts   = nom.split()
    cognome = parts[-1] if parts else nom
    nome    = " ".join(parts[:-1]) if len(parts) > 1 else ""

    email, conf = find_email(cognome, nome)
    anno = _parse_year(dcons)
    flags = []
    if not email:
        flags.append("NO_EMAIL")

    rec = {
        "fonte":        "Sheet3-Metodologiche",
        "cognome":      cognome,
        "nome":         nome,
        "email":        email,
        "email_conf":   conf,
        "cert_code":    num,
        "cert_name":    f"{tipo} — {desc}",
        "area":         "METODOLOGICHE",
        "cluster":      tipo,
        "anno":         anno,
        "note_orig":    f"Scadenza: {dscad}" if dscad else "",
        "status":       "",
        "pdf_file":     "",
        "pdf_cat":      "",
        "pdf_conf":     "",
        "flag":         " | ".join(flags),
        "note_rev":     "",
    }
    all_certs.append(rec)

print(f"      Cert totali (tutte le fonti): {len(all_certs)}")

# ── Canonicalizza descrizioni per codice ──
code_incoerenti: list[dict] = []
for code, descs in code_to_descs.items():
    # Scegli canonical = descrizione più frequente (a parità: più lunga)
    canonical = sorted(descs.items(), key=lambda x: (x[1], len(x[0])), reverse=True)[0][0]
    code_canonical[code] = canonical
    if len(descs) > 1:
        code_incoerenti.append({
            "cert_code":    code,
            "n_varianti":   len(descs),
            "canonica":     canonical,
            "varianti":     " | ".join(f'"{d}" ({c}x)' for d, c in descs.items()),
        })

# Applica canonical a tutti i record
for rec in all_certs:
    if rec["cert_code"] and rec["cert_code"] in code_canonical:
        original = rec["cert_name"]
        canonical = code_canonical[rec["cert_code"]]
        if _norm(original) != _norm(canonical):
            rec["cert_name"] = canonical
            if "NOME_UNIFORMATO" not in rec["flag"]:
                rec["flag"] = (rec["flag"] + " | NOME_UNIFORMATO").strip(" |")

print(f"      Codici con descrizioni incoerenti: {len(code_incoerenti)}")


# ──────────────────────────────────────────────────────────────────────────────
# STEP 5 — MATCH PDF → CERT RECORD
# ──────────────────────────────────────────────────────────────────────────────

print("[5/6] Match PDF → persone/certificazioni...")

# Build lookup: norm(cognome) → [indici in all_certs]
cog_idx: dict[str, list] = defaultdict(list)
for i, rec in enumerate(all_certs):
    cog_idx[_norm(rec["cognome"])].append(i)

pdf_no_match: list[dict] = []  # PDF senza match Excel
pdf_matched_ids: set = set()   # indici all_certs già matchati con un PDF

for pdf in pdf_records:
    surn = pdf["surname_pdf"]   # es. "MORISCO"
    cname_raw = pdf["cert_name_raw"]

    # Match cognome
    cand_indices: list[tuple[int, float]] = []  # (idx, score)
    surn_norm = _norm(surn)
    for cog_n, idxs in cog_idx.items():
        s = _sim(surn, cog_n)
        if s >= THRESH_SURNAME:
            for i in idxs:
                cand_indices.append((i, s))

    if not cand_indices:
        pdf_no_match.append({
            "file":           pdf["file"],
            "rel_path":       pdf["rel_path"],
            "category":       pdf["category"],
            "surname_pdf":    surn,
            "cert_name_raw":  cname_raw,
            "note":           "Nessun cognome corrispondente in Excel",
        })
        continue

    # Tra i candidati per cognome, match il nome cert
    best_idx   = -1
    best_score = -1.0
    for idx, _ in cand_indices:
        rec = all_certs[idx]
        cert_excel = rec["cert_name"] or rec["cert_code"]
        s = _sim(cname_raw, cert_excel)
        if s > best_score:
            best_score = s
            best_idx   = idx

    pdf["matched_idx"]   = best_idx
    pdf["match_score"]   = round(best_score, 3)
    pdf["surname_score"] = round(max(s for _, s in cand_indices), 3)

    rec = all_certs[best_idx]
    if best_score >= THRESH_CERTNAME:
        # Match buono
        conf_str = f"{best_score:.0%}"
        rec["pdf_file"] = pdf["file"]
        rec["pdf_cat"]  = pdf["category"]
        rec["pdf_conf"] = conf_str
        pdf_matched_ids.add(best_idx)
        if best_score < 0.85 and "PDF_MATCH_DUBBIO" not in rec["flag"]:
            rec["flag"] = (rec["flag"] + " | PDF_MATCH_DUBBIO").strip(" |")
    else:
        # Cognome trovato ma cert name non batte soglia
        conf_str = f"{best_score:.0%}"
        rec["flag"] = (rec["flag"] + " | PDF_CERT_DUBBIO").strip(" |")
        # Aggiungi comunque il file come possibilità
        if not rec["pdf_file"]:
            rec["pdf_file"] = pdf["file"] + " (?)"
            rec["pdf_cat"]  = pdf["category"]
            rec["pdf_conf"] = conf_str

# Cert senza PDF
for i, rec in enumerate(all_certs):
    if rec["fonte"] != "Sheet3-Metodologiche" and not rec["pdf_file"]:
        if "PDF_MANCANTE" not in rec["flag"] and "NO_CODICE" not in rec["flag"]:
            rec["flag"] = (rec["flag"] + " | PDF_MANCANTE").strip(" |")

# Duplicati (stesso email + cert_code)
seen_keys: set = set()
for rec in all_certs:
    k = (_norm(rec["email"]), _norm(rec["cert_code"]))
    if k in seen_keys and rec["email"] and rec["cert_code"]:
        rec["flag"] = (rec["flag"] + " | DUPLICATO").strip(" |")
    else:
        seen_keys.add(k)

print(f"      PDF senza match: {len(pdf_no_match)}")
print(f"      Cert senza PDF:  {sum(1 for r in all_certs if 'PDF_MANCANTE' in r['flag'])}")


# ──────────────────────────────────────────────────────────────────────────────
# STEP 5b — VERIFICA AI DEI PDF  (solo con --ai o --ai-all)
# ──────────────────────────────────────────────────────────────────────────────

# {cert_idx → ai_result dict}
ai_results: dict[int, dict] = {}

if RUN_AI:
    # Carica API key
    api_key = os.environ.get("OPENAI_API_KEY", "")
    if not api_key:
        env_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), ".env"
        )
        if os.path.exists(env_path):
            for line in open(env_path, encoding="utf-8"):
                if line.startswith("OPENAI_API_KEY="):
                    api_key = line.split("=", 1)[1].strip().strip('"').strip("'")
    if not api_key:
        print("[AI] ATTENZIONE: OPENAI_API_KEY non trovata — skip verifica AI")
        RUN_AI = False

if RUN_AI:
    import pymupdf                              # type: ignore
    from openai import OpenAI                   # type: ignore

    oai = OpenAI(api_key=api_key)

    # Seleziona PDF da verificare
    # --ai-all  → tutti i PDF abbinati (pdf_records con matched_idx)
    # --ai      → solo quelli con flag dubbio
    DUBBI_FLAGS = {"PDF_MATCH_DUBBIO", "PDF_CERT_DUBBIO"}

    to_verify: list[dict] = []
    for pdf in pdf_records:
        idx = pdf.get("matched_idx", -1)
        if idx < 0:
            continue
        rec   = all_certs[idx]
        flags = set(f.strip() for f in rec["flag"].split("|") if f.strip())
        if RUN_AI_ALL or (flags & DUBBI_FLAGS):
            to_verify.append(pdf)

    print(f"[5b] Verifica AI su {len(to_verify)} PDF "
          f"({'tutti abbinati' if RUN_AI_ALL else 'solo match dubbi'})...")

    def _extract_text(filepath: str) -> tuple[str, bool]:
        """Estrae testo dal PDF. Ritorna (testo, is_scanned)."""
        try:
            doc  = pymupdf.open(filepath)
            text = "\n".join(page.get_text() for page in doc).strip()
            doc.close()
            return text, len(text) < 80   # meno di 80 char → probabilmente scansionato
        except Exception as e:
            return f"[ERRORE ESTRAZIONE: {e}]", True

    def _extract_image_b64(filepath: str) -> str | None:
        """Prima pagina PDF come immagine base64 per vision (PDF scansionati)."""
        try:
            import base64
            doc  = pymupdf.open(filepath)
            page = doc[0]
            mat  = pymupdf.Matrix(1.5, 1.5)   # 108 DPI → buon compromesso qualità/costo
            pix  = page.get_pixmap(matrix=mat)
            doc.close()
            return base64.b64encode(pix.tobytes("png")).decode()
        except Exception:
            return None

    AI_SYSTEM = (
        "Sei un esperto di certificazioni SAP. "
        "Analizzi testi di certificati e rispondi ESCLUSIVAMENTE con JSON valido, "
        "senza markdown, senza testo aggiuntivo prima o dopo."
    )

    AI_PROMPT_TMPL = """\
Verifica questo certificato.

DATI ATTESI:
- Intestatario atteso: {cognome} {nome}
- Certificazione attesa: {cert_name}
- Codice atteso: {cert_code}
- Anno atteso: {anno}

TESTO ESTRATTO DAL PDF:
---
{text}
---

Rispondi con questo JSON (senza markdown):
{{
  "match_persona": "SI" | "NO" | "INCERTO",
  "match_cert":    "SI" | "NO" | "INCERTO",
  "persona_trovata": "<nome esatto trovato nel documento>",
  "cert_trovata":    "<nome certificazione trovato nel documento>",
  "anno_trovato":    "<anno trovato nel documento>",
  "confidenza":      0.0-1.0,
  "note":            "<eventuali anomalie o osservazioni brevi>"
}}"""

    def _verify_one(pdf: dict) -> tuple[int, dict]:
        """Verifica un singolo PDF con OpenAI. Ritorna (cert_idx, result)."""
        idx = pdf["matched_idx"]
        rec = all_certs[idx]
        fpath = os.path.join(CERT_DIR, pdf["rel_path"])

        text, is_scanned = _extract_text(fpath)

        result = {
            "file":          pdf["file"],
            "cert_idx":      idx,
            "cognome":       rec["cognome"],
            "nome":          rec["nome"],
            "cert_code":     rec["cert_code"],
            "cert_name":     rec["cert_name"],
            "is_scanned":    is_scanned,
            "text_chars":    len(text),
            "match_persona": "N/A",
            "match_cert":    "N/A",
            "persona_trovata": "",
            "cert_trovata":    "",
            "anno_trovato":    "",
            "confidenza":      "",
            "note":            "",
            "ai_model":        "",
            "errore":          "",
        }

        try:
            if is_scanned:
                # PDF scansionato → usa vision (GPT-4o)
                img_b64 = _extract_image_b64(fpath)
                if not img_b64:
                    result["errore"] = "Impossibile estrarre immagine"
                    return idx, result

                prompt_text = (
                    f"Verifica questo certificato.\n\n"
                    f"DATI ATTESI:\n"
                    f"- Intestatario: {rec['cognome']} {rec['nome']}\n"
                    f"- Certificazione: {rec['cert_name']}\n"
                    f"- Codice: {rec['cert_code']}\n"
                    f"- Anno: {rec['anno']}\n\n"
                    f"Rispondi SOLO con JSON:\n"
                    f'{{"match_persona":"SI"|"NO"|"INCERTO",'
                    f'"match_cert":"SI"|"NO"|"INCERTO",'
                    f'"persona_trovata":"...","cert_trovata":"...",'
                    f'"anno_trovato":"...","confidenza":0.0-1.0,"note":"..."}}'
                )
                resp = oai.chat.completions.create(
                    model="gpt-4o",
                    max_tokens=300,
                    temperature=0,
                    messages=[
                        {"role": "system", "content": AI_SYSTEM},
                        {"role": "user", "content": [
                            {"type": "text", "text": prompt_text},
                            {"type": "image_url", "image_url": {
                                "url": f"data:image/png;base64,{img_b64}",
                                "detail": "low",
                            }},
                        ]},
                    ],
                )
                result["ai_model"] = "gpt-4o (vision)"
            else:
                # PDF digitale → usa testo (GPT-4o-mini, più economico)
                # Tronca a 3000 char per risparmiare token
                text_trunc = text[:3000]
                prompt = AI_PROMPT_TMPL.format(
                    cognome   = rec["cognome"],
                    nome      = rec["nome"],
                    cert_name = rec["cert_name"],
                    cert_code = rec["cert_code"],
                    anno      = rec["anno"],
                    text      = text_trunc,
                )
                resp = oai.chat.completions.create(
                    model="gpt-4o-mini",
                    max_tokens=300,
                    temperature=0,
                    messages=[
                        {"role": "system", "content": AI_SYSTEM},
                        {"role": "user",   "content": prompt},
                    ],
                )
                result["ai_model"] = "gpt-4o-mini"

            raw = resp.choices[0].message.content.strip()
            # Rimuovi eventuale markdown residuo
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
            parsed = json.loads(raw)

            result.update({
                "match_persona":   parsed.get("match_persona", ""),
                "match_cert":      parsed.get("match_cert", ""),
                "persona_trovata": parsed.get("persona_trovata", ""),
                "cert_trovata":    parsed.get("cert_trovata", ""),
                "anno_trovato":    str(parsed.get("anno_trovato", "")),
                "confidenza":      str(parsed.get("confidenza", "")),
                "note":            parsed.get("note", ""),
            })

        except json.JSONDecodeError as e:
            result["errore"] = f"JSON non valido: {e} — risposta raw: {raw[:200]}"
        except Exception as e:
            result["errore"] = str(e)[:300]

        return idx, result

    # Esecuzione parallela (max 8 worker — rate limit OpenAI)
    MAX_WORKERS = 8
    done = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_verify_one, pdf): pdf for pdf in to_verify}
        for future in as_completed(futures):
            done += 1
            try:
                idx, result = future.result()
                ai_results[idx] = result
                mp = result["match_persona"]
                mc = result["match_cert"]
                status_icon = "✅" if mp == "SI" and mc == "SI" else (
                              "❌" if mp == "NO" or mc == "NO" else "⚠️")
                print(f"  [{done}/{len(to_verify)}] {status_icon} "
                      f"{result['cognome']} | {result['file'][:50]} "
                      f"→ persona={mp} cert={mc}")
            except Exception as e:
                print(f"  [{done}/{len(to_verify)}] ERRORE: {e}")

    # Aggiorna flag sui record
    for idx, res in ai_results.items():
        rec = all_certs[idx]
        if res["errore"]:
            _flag_add(rec, "AI_ERRORE")
        elif res["match_persona"] == "NO" or res["match_cert"] == "NO":
            _flag_add(rec, "AI_MISMATCH")
        elif res["match_persona"] == "SI" and res["match_cert"] == "SI":
            # Rimuovi flag dubbio se AI lo ha confermato
            rec["flag"] = re.sub(r"\s*\|\s*(PDF_MATCH_DUBBIO|PDF_CERT_DUBBIO)", "", rec["flag"]).strip(" |")
            _flag_add(rec, "AI_OK")

    n_ok      = sum(1 for r in ai_results.values() if r["match_persona"]=="SI" and r["match_cert"]=="SI")
    n_mismatch= sum(1 for r in ai_results.values() if r["match_persona"]=="NO" or r["match_cert"]=="NO")
    n_incerto = len(ai_results) - n_ok - n_mismatch
    print(f"      AI: {n_ok} OK | {n_mismatch} MISMATCH | {n_incerto} INCERTI")


# ──────────────────────────────────────────────────────────────────────────────
# STEP 6 — MASTERDATA: AREE E CLUSTER
# ──────────────────────────────────────────────────────────────────────────────

aree_set: dict[str, int] = defaultdict(int)
cluster_set: dict[str, dict] = defaultdict(lambda: {"area": "", "count": 0})

for rec in all_certs:
    if rec["area"]:
        aree_set[rec["area"]] += 1
    if rec["cluster"]:
        cluster_set[rec["cluster"]]["count"] += 1
        if not cluster_set[rec["cluster"]]["area"]:
            cluster_set[rec["cluster"]]["area"] = rec["area"]

# Aggiungi aree dai folder PDF
for pdf in pdf_records:
    cat = pdf["category"].replace("_", " ").title()
    aree_set[cat] += 0   # assicura che compaiano anche se senza cert Excel


# ──────────────────────────────────────────────────────────────────────────────
# STEP 7 — SCRIVI EXCEL OUTPUT
# ──────────────────────────────────────────────────────────────────────────────

print("[6/6] Scrittura Excel output...")

wb_out = openpyxl.Workbook()
del wb_out[wb_out.sheetnames[0]]

# ── Stili ──
H_FILL   = PatternFill("solid", fgColor="1F4E79")
H_FONT   = Font(bold=True, color="FFFFFF", size=10)
FLAG_FILL = {
    "NO_EMAIL":        PatternFill("solid", fgColor="FFD7D7"),  # rosso chiaro
    "PDF_MANCANTE":    PatternFill("solid", fgColor="FFF0CC"),  # giallo
    "PDF_MATCH_DUBBIO":PatternFill("solid", fgColor="FFE4B5"),  # arancio chiaro
    "PDF_CERT_DUBBIO": PatternFill("solid", fgColor="FFE4B5"),
    "CODICE_INCOER":   PatternFill("solid", fgColor="E8D5FF"),  # viola chiaro
    "OK":              PatternFill("solid", fgColor="E8F5E9"),   # verde chiaro
    "DUPLICATO":       PatternFill("solid", fgColor="FFD7D7"),
}
WRAP = Alignment(wrap_text=True, vertical="top")

def _add_sheet(name: str, headers: list, rows: list, col_widths: list = None) -> None:
    ws = wb_out.create_sheet(title=name)
    ws.freeze_panes = "A2"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(rows, start=2):
        ws.append(row)
        # colorazione per flag
        flag_col_idx = None
        for j, h in enumerate(headers):
            if h == "FLAG":
                flag_col_idx = j + 1
        if flag_col_idx:
            flag_val = ws.cell(i, flag_col_idx).value or ""
            fill = None
            if "NO_EMAIL" in flag_val or "DUPLICATO" in flag_val:
                fill = FLAG_FILL["NO_EMAIL"]
            elif "PDF_MANCANTE" in flag_val:
                fill = FLAG_FILL["PDF_MANCANTE"]
            elif "DUBBIO" in flag_val or "UNIFORMATO" in flag_val:
                fill = FLAG_FILL["PDF_MATCH_DUBBIO"]
            elif not flag_val.strip():
                fill = FLAG_FILL["OK"]
            if fill:
                for cell in ws[i]:
                    cell.fill = fill
        for cell in ws[i]:
            cell.alignment = WRAP
    # larghezze colonne
    if col_widths:
        for idx, w in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(1, idx).column_letter].width = w
    else:
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 55)


# ── Sheet 1: CERT_COMPLETO ──
cert_headers = [
    "FONTE", "COGNOME", "NOME", "EMAIL", "EMAIL_CONF",
    "CERT_CODE", "CERT_NAME", "AREA", "CLUSTER", "ANNO", "STATUS",
    "PDF_FILE", "PDF_CATEGORIA", "PDF_CONF",
    "AI_PERSONA", "AI_CERT", "AI_CONFIDENZA",   # colonne AI (vuote se --ai non usato)
    "NOTE_ORIG", "FLAG", "NOTE_REVISIONE"
]
cert_rows = []
sorted_certs = sorted(all_certs, key=lambda r: (_norm(r["cognome"]), _norm(r["cert_code"] or "")))
for i_orig, rec in enumerate(all_certs):
    # indice originale per lookup ai_results
    pass
# Ricostruiamo con indice originale
idx_sorted = sorted(range(len(all_certs)),
                    key=lambda i: (_norm(all_certs[i]["cognome"]),
                                   _norm(all_certs[i]["cert_code"] or "")))
for orig_idx in idx_sorted:
    rec = all_certs[orig_idx]
    ai  = ai_results.get(orig_idx, {})
    cert_rows.append([
        rec["fonte"], rec["cognome"], rec["nome"], rec["email"],
        f"{rec['email_conf']:.0%}" if rec["email_conf"] else "",
        rec["cert_code"], rec["cert_name"], rec["area"], rec["cluster"],
        rec["anno"], rec["status"],
        rec["pdf_file"], rec["pdf_cat"], rec["pdf_conf"],
        ai.get("match_persona", ""),
        ai.get("match_cert", ""),
        ai.get("confidenza", ""),
        rec["note_orig"], rec["flag"], rec["note_rev"],
    ])

_add_sheet("CERT_COMPLETO", cert_headers, cert_rows,
           [12, 14, 14, 32, 10, 16, 55, 15, 20, 8, 12, 45, 22, 10, 10, 10, 10, 25, 35, 30])


# ── Sheet 2: PERSONE_SENZA_EMAIL ──
ne_headers = ["COGNOME", "NOME", "N_CERT", "POSSIBILI_EMAIL", "NOTE_REVISIONE"]
ne_seen = {}
for rec in all_certs:
    if "NO_EMAIL" in rec["flag"]:
        k = (_norm(rec["cognome"]), _norm(rec["nome"]))
        if k not in ne_seen:
            ne_seen[k] = {"cognome": rec["cognome"], "nome": rec["nome"], "count": 0, "suggest": ""}
        ne_seen[k]["count"] += 1

# Suggerisci email plausibile: nome.cognome@mashfrog.com
for k, v in ne_seen.items():
    first = _norm(v["nome"]).split()[0] if v["nome"] else ""
    last  = _norm(v["cognome"])
    # Converti da unicode normalizzato
    sug = f"{first}.{last}@mashfrog.com".replace(" ", ".")
    v["suggest"] = sug

ne_rows = [[v["cognome"], v["nome"], v["count"], v["suggest"], ""] for v in ne_seen.values()]
ne_rows.sort(key=lambda r: r[0])
_add_sheet("PERSONE_SENZA_EMAIL", ne_headers, ne_rows, [16, 16, 8, 40, 40])


# ── Sheet 3: PDF_SENZA_MATCH ──
pm_headers = ["FILE", "REL_PATH", "CATEGORIA", "COGNOME_PDF", "CERT_NAME_ESTRATTO",
              "COGNOME_PIU_SIMILE", "SCORE", "NOTE_REVISIONE"]
pm_rows = []
for pdf in pdf_no_match:
    # Cerca cognome più simile in Sheet2 (informativo)
    best_sn = ""; best_sc = 0.0
    for cog_n in cog_idx:
        s = _sim(pdf["surname_pdf"], cog_n)
        if s > best_sc:
            best_sc = s; best_sn = cog_n
    pm_rows.append([
        pdf["file"], pdf["rel_path"], pdf["category"],
        pdf["surname_pdf"], pdf["cert_name_raw"],
        best_sn if best_sc >= 0.6 else "", f"{best_sc:.0%}" if best_sc >= 0.6 else "",
        "",
    ])
_add_sheet("PDF_SENZA_MATCH", pm_headers, pm_rows, [45, 65, 25, 16, 55, 20, 8, 35])


# ── Sheet 4: CODICI_INCOERENTI ──
ci_headers = ["CERT_CODE", "N_VARIANTI", "NOME_CANONICO_SCELTO", "TUTTE_LE_VARIANTI", "APPROVARE? (SI/NO)"]
ci_rows = sorted(
    [[r["cert_code"], r["n_varianti"], r["canonica"], r["varianti"], ""] for r in code_incoerenti],
    key=lambda r: -r[1]
)
_add_sheet("CODICI_INCOERENTI", ci_headers, ci_rows, [18, 10, 70, 80, 15])


# ── Sheet 5: AREE_MASTERDATA ──
ar_headers = ["AREA", "N_CERT", "DESCRIZIONE", "AREA_FOLDER_PDF"]
ar_rows = sorted([[a, c, "", ""] for a, c in aree_set.items() if a], key=lambda r: r[0])
_add_sheet("AREE_MASTERDATA", ar_headers, ar_rows, [25, 10, 50, 30])


# ── Sheet 6: CLUSTER_MASTERDATA ──
cl_headers = ["CLUSTER", "AREA", "N_CERT", "DESCRIZIONE"]
cl_rows = sorted(
    [[cl, v["area"], v["count"], ""] for cl, v in cluster_set.items() if cl],
    key=lambda r: r[0]
)
_add_sheet("CLUSTER_MASTERDATA", cl_headers, cl_rows, [30, 20, 10, 50])


# ── Sheet 7: AI_VERIFICA (solo se --ai eseguito) ──
if ai_results:
    av_headers = [
        "FILE", "COGNOME", "NOME", "CERT_CODE", "CERT_NAME",
        "SCANSIONATO", "CHARS_ESTRATTI", "MODELLO_AI",
        "MATCH_PERSONA", "MATCH_CERT", "CONFIDENZA",
        "PERSONA_TROVATA", "CERT_TROVATA", "ANNO_TROVATO",
        "NOTE_AI", "ERRORE"
    ]
    av_rows = []
    for res in sorted(ai_results.values(), key=lambda r: r["cognome"]):
        av_rows.append([
            res["file"], res["cognome"], res["nome"],
            res["cert_code"], res["cert_name"],
            "SI" if res["is_scanned"] else "NO",
            res["text_chars"], res["ai_model"],
            res["match_persona"], res["match_cert"], res["confidenza"],
            res["persona_trovata"], res["cert_trovata"], res["anno_trovato"],
            res["note"], res["errore"],
        ])
    _add_sheet("AI_VERIFICA", av_headers, av_rows,
               [45, 14, 14, 16, 55, 10, 12, 18, 12, 12, 12, 25, 55, 12, 40, 40])

# ── SALVA ──
wb_out.save(XLS_OUT)

print()
print("=" * 60)
print(f"  OUTPUT: {XLS_OUT}")
print("=" * 60)
print(f"  Sheet CERT_COMPLETO:      {len(cert_rows)} righe")
print(f"  Sheet PERSONE_SENZA_EMAIL:{len(ne_rows)} persone")
print(f"  Sheet PDF_SENZA_MATCH:    {len(pm_rows)} PDF")
print(f"  Sheet CODICI_INCOERENTI:  {len(ci_rows)} codici")
print(f"  Sheet AREE_MASTERDATA:    {len(ar_rows)} aree")
print(f"  Sheet CLUSTER_MASTERDATA: {len(cl_rows)} cluster")
if ai_results:
    n_ok  = sum(1 for r in ai_results.values() if r["match_persona"]=="SI" and r["match_cert"]=="SI")
    n_no  = sum(1 for r in ai_results.values() if r["match_persona"]=="NO" or r["match_cert"]=="NO")
    n_inc = len(ai_results) - n_ok - n_no
    print(f"  Sheet AI_VERIFICA:        {len(ai_results)} PDF verificati "
          f"({n_ok} OK / {n_no} MISMATCH / {n_inc} INCERTI)")
print()
print("  FLAGS presenti in CERT_COMPLETO:")
from collections import Counter
flag_counts = Counter()
for rec in all_certs:
    for f in rec["flag"].split("|"):
        f = f.strip()
        if f:
            flag_counts[f] += 1
for flag, cnt in sorted(flag_counts.items(), key=lambda x: -x[1]):
    print(f"    {flag:30s} {cnt}")
print()
print("  Pronto per revisione manuale. Una volta pulito riesegui con --generate-store.")
